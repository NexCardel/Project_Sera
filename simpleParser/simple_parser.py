"""Simple, evidence-first parser for ``seraRawPayloadDump.txt``.

This pipeline intentionally favors explicit fields, endpoint context, and
visible ambiguity over aggressive inference. It is independent from the
legacy FST classifier and the evidence-first tracer so its output can be used
as a cross-check.
"""

from __future__ import annotations

import json
import base64
import binascii
import os
import re
import time
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import parse_qs, urlparse

PAN_RE = re.compile(r"^[A-Z]{5}[0-9]{4}[A-Z]$")
PAN_SEARCH_RE = re.compile(r"\b[A-Z]{5}[0-9]{4}[A-Z]\b", re.I)
GSTIN_RE = re.compile(r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][1-9A-Z]Z[0-9A-Z]$")
GSTIN_SEARCH_RE = re.compile(r"\b[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][1-9A-Z]Z[0-9A-Z]\b", re.I)
EMAIL_RE = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")
ENTRY_RE = re.compile(r"CAPTURE DUMP ENTRY #\s*(\d+)")
SEPARATOR = "========================================================================================"

IDENTITY_KEYS = {
    "pan", "pannumber", "pan_no", "entitynum", "submituserid", "loggedinuserid",
    "userid", "user_id", "gstin", "gstinnumber", "assesseverpan",
}
NAME_KEYS = {
    "assesseevername", "auth_name", "fullname", "nameasperbank", "bn", "ln", "tn",
    "firstname", "middlename", "midname", "lastname", "surnameororgname",
}
EMAIL_KEYS = {"email", "emailid", "priemailid", "emailaddress"}
PHONE_KEYS = {"mobileno", "mobile", "mobilenumber", "phone", "phonenumber", "primobilenum", "primarymobile"}
DOB_KEYS = {"dob", "dateofbirth"}
TRANSACTION_KEYS = {
    "arn", "arnnumber", "acknum", "transactionno", "uniquereqid", "reqid",
    "itbasequenceno", "ref_id", "referenceid", "commrefno", "receipt",
}
FILING_ACTIONS = {"return submission", "gst return filing"}


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    value = re.sub(r"\s+", " ", str(value)).strip()
    return value or None


def _decode_field_value(key: str, value: Any) -> str | None:
    """Decode known Base64-obfuscated contact/name fields only.

    We do not decode arbitrary strings: identifiers and transaction values can
    coincidentally look like Base64 and must remain unchanged.
    """
    text = _clean(value)
    if not text:
        return text
    key_l = key.lower().replace("_", "")
    known = EMAIL_KEYS | PHONE_KEYS | NAME_KEYS
    if key_l not in {item.replace("_", "") for item in known}:
        return text
    if not re.fullmatch(r"[A-Za-z0-9+/]+={0,2}", text) or len(text) % 4:
        return text
    try:
        decoded = base64.b64decode(text, validate=True).decode("utf-8").strip()
    except (binascii.Error, UnicodeDecodeError, ValueError):
        return text
    if key_l in {item.replace("_", "") for item in EMAIL_KEYS}:
        return decoded if EMAIL_RE.fullmatch(decoded) else text
    if key_l in {item.replace("_", "") for item in PHONE_KEYS}:
        return decoded if re.fullmatch(r"\+?[0-9][0-9 .-]{6,18}", decoded) else text
    if key_l in {item.replace("_", "") for item in NAME_KEYS}:
        return decoded if re.fullmatch(r"[A-Za-z][A-Za-z .,'-]{1,120}", decoded) else text
    return text


def _valid_pan(value: Any) -> str | None:
    value = _clean(value)
    if not value:
        return None
    value = value.upper()
    return value if PAN_RE.fullmatch(value) else None


def _valid_gstin(value: Any) -> str | None:
    value = _clean(value)
    if not value:
        return None
    value = value.upper()
    return value if GSTIN_RE.fullmatch(value) else None


def _parse_json_string(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not (text.startswith("{") or text.startswith("[")):
        return value
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return value


def _walk(value: Any, path: str = "") -> Iterable[tuple[str, str, Any]]:
    if isinstance(value, dict):
        for key, child in value.items():
            child_path = f"{path}.{key}" if path else str(key)
            yield child_path, str(key), child
            decoded = _parse_json_string(child)
            if decoded is not child:
                yield from _walk(decoded, f"{child_path} (decoded)")
            elif isinstance(child, (dict, list)):
                yield from _walk(child, child_path)
    elif isinstance(value, list):
        for index, child in enumerate(value):
            child_path = f"{path}[{index}]"
            decoded = _parse_json_string(child)
            if decoded is not child:
                yield from _walk(decoded, f"{child_path} (decoded)")
            elif isinstance(child, (dict, list)):
                yield from _walk(child, child_path)


def _header(chunk: str) -> dict[str, Any]:
    result: dict[str, Any] = {}
    match = ENTRY_RE.search(chunk)
    result["entry_number"] = int(match.group(1)) if match else None
    for line in chunk.splitlines():
        if ":" not in line or line.strip().startswith("RAW JSON"):
            continue
        key, value = line.split(":", 1)
        result[key.strip().lower().replace(" ", "_").replace("/", "_")] = value.strip()
    return result


def parse_dump(source: str | Path) -> list[dict[str, Any]]:
    path = Path(source)
    text = path.read_text(encoding="utf-8", errors="replace") if path.is_file() else str(source)
    entries: list[dict[str, Any]] = []
    for chunk in re.split(r"(?=CAPTURE DUMP ENTRY #)", text):
        if not ENTRY_RE.search(chunk):
            continue
        item = _header(chunk)
        body = chunk.split("RAW JSON PAYLOAD:", 1)[1] if "RAW JSON PAYLOAD:" in chunk else ""
        body = body.split(SEPARATOR, 1)[0].strip()
        try:
            item["payload"] = json.loads(body)
            item["parse_error"] = None
        except json.JSONDecodeError as exc:
            item["payload"] = None
            item["parse_error"] = f"JSON decode error: {exc.msg} at character {exc.pos}"
        entries.append(item)
    return entries


def _all_values(item: dict[str, Any]) -> list[tuple[str, str, Any]]:
    payload = item.get("payload") or {}
    return list(_walk(payload))


def _epoch_date(value: Any) -> str | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    # The dump uses epoch milliseconds. Avoid accepting seconds or arbitrary IDs.
    if abs(number) < 10_000_000_000 or abs(number) > 10_000_000_000_000:
        return None
    try:
        converted = datetime(1970, 1, 1, tzinfo=timezone.utc) + timedelta(milliseconds=number)
        return converted.date().isoformat()
    except (OverflowError, ValueError):
        return None


def _date_text(value: Any) -> str | None:
    value = _clean(value)
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(value, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _valid_dob(value: str | None) -> str | None:
    if not value:
        return None
    try:
        parsed = date.fromisoformat(value)
    except ValueError:
        return None
    today = date.today()
    return value if date(1900, 1, 1) <= parsed <= today else None


def _normalize_return_type(url: str, values: list[tuple[str, str, Any]], envelope: dict[str, Any]) -> tuple[str, str] | tuple[None, None]:
    query = parse_qs(urlparse(url).query)
    rtn = _clean(query.get("rtn_typ", [None])[0])
    if rtn:
        return (rtn.upper().replace("GSTR1", "GSTR-1"), "url.rtn_typ")
    for path, key, value in values:
        key_l = key.lower()
        if key_l == "formname" and isinstance(value, str) and value.strip().lower() not in {"", "none", "null"}:
            text = _clean(value)
            if text and not text.upper().startswith("FO-"):
                return (text.upper(), path)
        if key_l == "form" and "gstr1iff" in path.lower() and _clean(value):
            return (_clean(value).upper(), path)
    for key in ("filing_type", "form_type_cd", "form_cd"):
        value = _clean(envelope.get(key))
        if not value or value.lower() in {"profile info", "gst taxpayer profile", "profile / contact details"}:
            continue
        if key in {"form_type_cd", "form_cd"} and ("eportal.incometax.gov.in" not in url.lower()):
            continue
        return ((f"ITR-{value}" if value.isdigit() else value.upper()), f"envelope.{key}")
    return (None, None)


def _period(url: str, item: dict[str, Any], values: list[tuple[str, str, Any]]) -> tuple[str, str, str] | tuple[None, None, None]:
    query = parse_qs(urlparse(url).query)
    rtn_prd = _clean(query.get("rtn_prd", [None])[0])
    if rtn_prd and re.fullmatch(r"(?:0[1-9]|1[0-2])[0-9]{4}", rtn_prd):
        return (f"{rtn_prd[2:]}-{rtn_prd[:2]}", "GST month", "url.rtn_prd")
    raw_label = _clean(item.get("period_label"))
    if raw_label and raw_label.lower() not in {"profile info", "n/a", "na"}:
        return (raw_label, "source period label", "header.period_label")
    for path, key, value in values:
        key_l = key.lower()
        text = _clean(value)
        if not text:
            continue
        if key_l in {"assessmn t yr", "assessmn tyr", "assmentyear", "assessmntyr", "assessmentyear"} and str(text).isdigit():
            year = int(text)
            return (f"AY {year}-{str(year + 1)[-2:]}", "assessment year", path)
    return (None, None, None)


def _event(item: dict[str, Any]) -> dict[str, Any]:
    envelope = item.get("payload") if isinstance(item.get("payload"), dict) else {}
    raw = envelope.get("raw_payload") if isinstance(envelope.get("raw_payload"), dict) else {}
    values = _all_values(item)
    pans: set[str] = set()
    gstins: set[str] = set()
    identity_sources: list[str] = []
    for path, key, value in values:
        key_l = key.lower()
        if key_l in IDENTITY_KEYS:
            gstin = _valid_gstin(value)
            pan = _valid_pan(value)
            if gstin:
                gstins.add(gstin); identity_sources.append(path)
                pans.add(gstin[2:12])
            elif pan:
                pans.add(pan); identity_sources.append(path)
    if not pans and not gstins:
        text = json.dumps(envelope, ensure_ascii=False)
        for gstin in GSTIN_SEARCH_RE.findall(text.upper()):
            gstins.add(gstin); pans.add(gstin[2:12]); identity_sources.append("fallback.GSTIN")
        for pan in PAN_SEARCH_RE.findall(text.upper()):
            pans.add(pan); identity_sources.append("fallback.PAN")

    names: list[dict[str, str]] = []
    emails: list[dict[str, str]] = []
    phones: list[dict[str, str]] = []
    dobs: list[dict[str, str]] = []
    transactions: list[dict[str, str]] = []
    components: dict[str, str] = {}
    for path, key, value in values:
        key_l = key.lower()
        text = _decode_field_value(key_l, value)
        if key_l in NAME_KEYS and text and text.upper() not in {"N/A", "NA", "NONE", "NULL"}:
            role = "business_name" if key_l in {"bn", "tn"} else ("legal_name" if key_l == "ln" else "client_name")
            names.append({"value": text, "source": path, "role": role})
        if key_l in {"firstname", "middlename", "midname", "lastname", "surnameororgname"} and text:
            components[key_l] = text
        if key_l in EMAIL_KEYS and text and EMAIL_RE.fullmatch(text):
            emails.append({"value": text, "source": path, "role": "primary_email" if key_l == "priemailid" else "portal_email"})
        if key_l in PHONE_KEYS and text and re.fullmatch(r"\+?[0-9][0-9 .-]{6,18}", text):
            phones.append({"value": text, "source": path, "role": "primary_mobile" if key_l in {"primobilenum", "primarymobile"} else "portal_mobile"})
        if key_l in DOB_KEYS:
            parsed = _valid_dob(_date_text(value) or _epoch_date(value))
            if parsed:
                dobs.append({"value": parsed, "source": path})
        if key_l in TRANSACTION_KEYS and text and text.lower() not in {"n/a", "na", "none", "null", "-"}:
            role = "everification_transaction" if text.upper().startswith("EVERIFY") else ("acknowledgement" if key_l in {"arn", "arnnumber", "acknum"} else "portal_transaction")
            transactions.append({"value": text, "source": path, "role": role})
    if components:
        ordered = [components.get(k) for k in ("firstname", "midname", "middlename", "lastname", "surnameororgname")]
        combined = " ".join(dict.fromkeys(x for x in ordered if x))
        if combined:
            names.append({"value": combined, "source": "composed.name_components", "role": "client_name"})

    url = _clean(envelope.get("url")) or ""
    return_type, return_source = _normalize_return_type(url, values, envelope)
    period, period_type, period_source = _period(url, item, values)
    # Prefer the extension's explicit evidence envelope when available. The
    # raw payload remains the fallback/source of record for older captures.
    extension_evidence = envelope.get("simple_parser_evidence") if isinstance(envelope.get("simple_parser_evidence"), dict) else {}
    identities = extension_evidence.get("identities") if isinstance(extension_evidence.get("identities"), dict) else {}
    for value in identities.get("pans", []):
        normalized = _valid_pan(value)
        if normalized:
            pans.add(normalized)
    for value in identities.get("gstins", []):
        normalized = _valid_gstin(value)
        if normalized:
            gstins.add(normalized); pans.add(normalized[2:12])
    for key, target in (("names", names), ("emails", emails), ("dobs", dobs), ("phones", phones), ("transactions", transactions)):
        for entry in extension_evidence.get(key, []):
            if isinstance(entry, dict) and _clean(entry.get("value")):
                candidate = {k: _clean(v) for k, v in entry.items() if _clean(v)}
                if candidate and not any(candidate.get("value") == old.get("value") and candidate.get("source") == old.get("source") for old in target):
                    target.append(candidate)
    evidence_return = extension_evidence.get("return_type")
    if not return_type and isinstance(evidence_return, dict) and _clean(evidence_return.get("value")):
        return_type, return_source = _clean(evidence_return["value"]), _clean(evidence_return.get("source"))
    evidence_period = extension_evidence.get("period")
    if not period and isinstance(evidence_period, dict) and _clean(evidence_period.get("value")):
        period, period_type, period_source = _clean(evidence_period["value"]), "extension evidence", _clean(evidence_period.get("source"))
    raw_text = json.dumps(envelope, ensure_ascii=False).lower()
    lower_url = url.lower()
    lifecycle = extension_evidence.get("lifecycle") if isinstance(extension_evidence.get("lifecycle"), dict) else {}
    submission = bool(lifecycle.get("submission"))
    if "/returns/submit/wzrd" in lower_url:
        submission = str(raw.get("httpStatus", "")).upper() == "ACCEPTED" and raw.get("successFlag") is True and any(t["role"] in {"portal_transaction", "acknowledgement"} for t in transactions)
    if "formdetails" in lower_url:
        submission = submission or '"status": "fil"' in raw_text
    everification = bool(lifecycle.get("everification")) or ("validateotp" in lower_url and str(raw.get("moduleCode", "")).upper() == "ITR" and str(raw.get("status", "")).upper() == "SUCCESS" and "otp validated" in raw_text)
    other_evc = bool(lifecycle.get("other_evc")) or (("validateotp" in lower_url and not everification and str(raw.get("status", "")).upper() == "SUCCESS") or "evc accepted" in raw_text)
    pending = bool(lifecycle.get("pending_everification")) or "pending for e-verification" in raw_text
    return {
        "entry": item.get("entry_number"), "timestamp": item.get("timestamp", ""), "portal": item.get("portal", ""),
        "url": url, "pans": sorted(pans), "gstins": sorted(gstins), "identity_sources": identity_sources,
        "names": names, "emails": emails, "phones": phones, "dobs": dobs, "transactions": transactions,
        "return_type": return_type, "return_type_source": return_source, "period": period,
        "period_type": period_type, "period_source": period_source, "submission": submission,
        "everification": everification, "other_evc": other_evc, "pending_everification": pending,
        "parse_error": item.get("parse_error"), "raw": envelope,
    }


def _refs(event: dict[str, Any]) -> set[str]:
    return {t["value"] for t in event["transactions"] if t["role"] == "acknowledgement"}


def _resolve_transaction_identities(events: list[dict[str, Any]]) -> None:
    """Propagate a PAN across payloads linked by a unique acknowledgement."""
    reference_to_pans: dict[str, set[str]] = defaultdict(set)
    for event in events:
        for reference in {t["value"] for t in event["transactions"]}:
            reference_to_pans[reference].update(event["pans"])
    for event in events:
        if event["pans"]:
            continue
        linked = set()
        for reference in {t["value"] for t in event["transactions"]}:
            linked.update(reference_to_pans.get(reference, set()))
        if len(linked) == 1:
            event["pans"] = sorted(linked)
            event["identity_sources"].append("transaction-link")


def _lifecycle_rows(events: list[dict[str, Any]], master_pans: Iterable[str] | None) -> tuple[list[list[Any]], list[dict[str, Any]]]:
    master = {str(p).strip().upper() for p in (master_pans or []) if PAN_RE.fullmatch(str(p).strip().upper())}
    by_pan: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for event in events:
        for pan in event["pans"]:
            by_pan[pan].append(event)
    rows: list[list[Any]] = []
    quarantine: list[dict[str, Any]] = []
    for pan, group in sorted(by_pan.items()):
        submissions = [e for e in group if e["submission"]]
        itr_evc = [e for e in group if e["everification"]]
        other_evc = [e for e in group if e["other_evc"]]
        linked = []
        for ev in itr_evc:
            if any(_refs(ev) & _refs(sub) for sub in submissions):
                linked.append(ev); continue
            if not _refs(ev) and len(submissions) == 1:
                linked.append(ev)
        if submissions and linked:
            category = "Return submitted and e-verified"
        elif submissions:
            category = "Return submitted — not e-verified"
        elif itr_evc:
            category = "E-verified only"
        elif other_evc:
            category = "EVC — no return submission"
        elif any(e["return_type"] or e["period"] for e in group):
            category = "Not submitted return"
        else:
            category = "No return activity observed"
        names = []
        emails = []
        phones = []
        dobs = []
        for e in group:
            names.extend(n["value"] for n in e["names"] if n["role"] != "business_name")
            emails.extend(x["value"] for x in e["emails"])
            phones.extend(x["value"] for x in e["phones"])
            dobs.extend(x["value"] for x in e["dobs"])
        rows.append([
            pan, category, "; ".join(dict.fromkeys(names)), "; ".join(dict.fromkeys(emails)), "; ".join(dict.fromkeys(dobs)),
            "; ".join(dict.fromkeys(e["return_type"] for e in group if e["return_type"])),
            "; ".join(dict.fromkeys(e["period"] for e in group if e["period"])),
            "; ".join(dict.fromkeys(t["value"] for e in submissions for t in e["transactions"])),
            "; ".join(dict.fromkeys(t["value"] for e in linked for t in e["transactions"])),
            "Yes" if pan in master else "No", "; ".join(dict.fromkeys(phones)),
        ])
    for e in events:
        if not e["pans"]:
            quarantine.append({"entry": e["entry"], "status": "unresolved_identity", "reason": "No defensible PAN/GSTIN candidate", "url": e["url"]})
        elif len(e["pans"]) > 1:
            quarantine.append({"entry": e["entry"], "status": "conflicting_identity", "reason": "Multiple PAN candidates in one payload", "url": e["url"]})
    return rows, quarantine


def _tracker_grouped_rows(events: list[dict[str, Any]]) -> tuple[list[str], list[list[Any]], list[str], list[list[Any]]]:
    """Build grouped tracker views with newest-first and oldest-first stacking."""
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for event in events:
        pans = event.get("pans") or ["UNRESOLVED"]
        for pan in pans:
            grouped[str(pan)].append(event)

    def event_sort_key(event: dict[str, Any]) -> tuple[str, int]:
        return (str(event.get("timestamp") or ""), int(event.get("entry") or 0))

    # Keep client groups together, but order the groups by their activity so
    # recent tracker captures are visible at the top of the UI-style view.
    ordered_groups_up = sorted(
        grouped.items(),
        key=lambda item: max((event_sort_key(event) for event in item[1]), default=("", 0)),
        reverse=True,
    )
    ordered_groups_down = sorted(
        grouped.items(),
        key=lambda item: min((event_sort_key(event) for event in item[1]), default=("", 0)),
    )
    event_headers = [
        "Entry", "Timestamp", "Portal", "Return Type", "Period", "Submission", "Transactions", "Endpoint"
    ]
    down_headers = ["PAN", "Client Names", "Business Names"] + event_headers
    down_rows: list[list[Any]] = []
    for pan, group in ordered_groups_down:
        ordered_events = sorted(group, key=event_sort_key)
        client_names = sorted({
            n["value"] for event in ordered_events for n in event.get("names", [])
            if n.get("role") != "business_name" and n.get("value")
        })
        business_names = sorted({
            n["value"] for event in ordered_events for n in event.get("names", [])
            if n.get("role") == "business_name" and n.get("value")
        })
        client_text = "; ".join(client_names)
        business_text = "; ".join(business_names)
        for event in ordered_events:
            down_rows.append([
                pan, client_text, business_text,
                event.get("entry"), event.get("timestamp", ""), event.get("portal", ""),
                event.get("return_type") or "", event.get("period") or "",
                "Yes" if event.get("submission") else "No",
                "; ".join(t["value"] for t in event.get("transactions", [])),
                event.get("url", ""),
            ])

    # Up-stacked mirrors the tracker UI: the newest entry is placed above
    # older entries, so a newly captured event appears at the top of its group.
    up_headers = down_headers[:]
    up_rows: list[list[Any]] = []
    for pan, group in ordered_groups_up:
        ordered_events = sorted(group, key=event_sort_key, reverse=True)
        client_names = sorted({
            n["value"] for event in group for n in event.get("names", [])
            if n.get("role") != "business_name" and n.get("value")
        })
        business_names = sorted({
            n["value"] for event in group for n in event.get("names", [])
            if n.get("role") == "business_name" and n.get("value")
        })
        for event in ordered_events:
            up_rows.append([
                pan, "; ".join(client_names), "; ".join(business_names),
                event.get("entry"), event.get("timestamp", ""), event.get("portal", ""),
                event.get("return_type") or "", event.get("period") or "",
                "Yes" if event.get("submission") else "No",
                "; ".join(t["value"] for t in event.get("transactions", [])),
                event.get("url", ""),
            ])
    return up_headers, up_rows, down_headers, down_rows


def _write_xlsx(result: dict[str, Any], output: str | Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    def add_sheet(name: str, headers: list[str], rows: Iterable[list[Any]]) -> None:
        ws = wb.create_sheet(name)
        ws.append(headers)
        for cell in ws[1]:
            cell.fill, cell.font = header_fill, header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in rows:
            ws.append(row)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for column in ws.columns:
            letter = column[0].column_letter
            width = max(len(str(c.value or "")) for c in column) + 2
            ws.column_dimensions[letter].width = min(max(width, 12), 55)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    add_sheet("Lifecycle Summary", ["PAN", "Lifecycle Category", "Client Names", "Emails", "DOBs", "Return Types", "Return Periods", "Submit Transactions", "E-Verification Transactions", "PAN Present in Master DB", "Mobile Numbers"], result["lifecycle_rows"])
    up_headers, up_rows, down_headers, down_rows = _tracker_grouped_rows(result["events"])
    add_sheet("trackerDumpGroupedUpStacked", up_headers, up_rows)
    add_sheet("trackerDumpGroupedDownStacked", down_headers, down_rows)
    add_sheet("Parser Events", ["Entry", "Timestamp", "PAN(s)", "Return Type", "Period", "Submission", "ITR E-Verification", "Other EVC", "Pending E-Verification", "Transactions", "Endpoint"], [
        [e["entry"], e["timestamp"], "; ".join(e["pans"]), e["return_type"] or "", e["period"] or "", "Yes" if e["submission"] else "No", "Yes" if e["everification"] else "No", "Yes" if e["other_evc"] else "No", "Yes" if e["pending_everification"] else "No", "; ".join(t["value"] for t in e["transactions"]), e["url"]]
        for e in result["events"]
    ])
    add_sheet("Identity Evidence", ["Entry", "PAN", "GSTIN", "Name", "Name Source", "Email", "Email Source", "Mobile", "Mobile Source", "DOB", "DOB Source", "Identity Source"], [
        [e["entry"], "; ".join(e["pans"]), "; ".join(e["gstins"]), "; ".join(n["value"] for n in e["names"]), "; ".join(n["source"] for n in e["names"]), "; ".join(x["value"] for x in e["emails"]), "; ".join(x["source"] for x in e["emails"]), "; ".join(x["value"] for x in e["phones"]), "; ".join(x["source"] for x in e["phones"]), "; ".join(x["value"] for x in e["dobs"]), "; ".join(x["source"] for x in e["dobs"]), "; ".join(e["identity_sources"])]
        for e in result["events"]
    ])
    add_sheet("Quarantine", ["Entry", "Status", "Reason", "Endpoint"], [[q["entry"], q["status"], q["reason"], q["url"]] for q in result["quarantine"]])

    output = Path(output)
    output.parent.mkdir(parents=True, exist_ok=True)
    temp = output.with_name(f".{output.stem}.{os.getpid()}.tmp.xlsx")
    wb.save(temp)
    try:
        wb.close()
    except Exception:
        pass
    try:
        os.replace(temp, output)
        return output
    except PermissionError:
        latest = output.with_name(f"{output.stem}_latest{output.suffix}")
        try:
            os.replace(temp, latest)
            return latest
        except PermissionError:
            fallback = output.with_name(f"{output.stem}_{int(time.time())}{output.suffix}")
            os.replace(temp, fallback)
            return fallback


def process_dump(input_dump: str | Path, output_excel: str | Path | None = None, master_pans: Iterable[str] | None = None) -> dict[str, Any]:
    events = [_event(item) for item in parse_dump(input_dump)]
    _resolve_transaction_identities(events)
    lifecycle_rows, quarantine = _lifecycle_rows(events, master_pans)
    result = {"events": events, "lifecycle_rows": lifecycle_rows, "quarantine": quarantine, "stats": {"input_entries": len(events), "lifecycle_rows": len(lifecycle_rows), "quarantine_entries": len(quarantine)}, "outputs": {}}
    if output_excel:
        result["outputs"]["excel_path"] = str(_write_xlsx(result, output_excel))
    return result


def main() -> None:
    import argparse
    parser = argparse.ArgumentParser(description="Project Sera Simple Parser")
    parser.add_argument("input_dump", nargs="?", default="..\\seraRawPayloadDump.txt")
    parser.add_argument("output_excel", nargs="?", default="simple_parser_report.xlsx")
    args = parser.parse_args()
    print(json.dumps(process_dump(args.input_dump, args.output_excel)["stats"], sort_keys=True))


if __name__ == "__main__":
    main()
