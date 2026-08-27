import os
import sys
import re
import json
import time
import argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- STYLING CONSTANTS ---
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F4E78")
SUB_TITLE_FONT = Font(name="Calibri", size=11, italic=True, color="595959")
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
)

COLORS = {
    "cat1": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"), # 1. Submitted Unverified (Yellow)
    "cat2": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), # 2. Submitted & Verified (Green)
    "cat3": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"), # 3. Bank Verified No Return (Blue)
    "cat4": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), # 4. GST Filed (Green)
    "bank_valid": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), # Bank Validated (Green)
    "bank_warn": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),  # Bank Valid with Warning (Yellow)
    "bank_failed": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"), # Bank Failed/Rejected (Red)
    "bank_disabled": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"), # Bank Disabled/Legacy (Gray)
    "cat6": PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid"), # 6. Visited No Return (Gray)
    "cat7": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"), # 7. Visited Site (All)
}

def resolve_target_dump(path):
    """Auto-detects live dump if path is generic or missing."""
    if os.path.exists(path):
        return path
    candidates = [
        os.path.join("..", "seraRawPayloadDump.txt"),
        "seraRawPayloadDump.txt",
        os.path.join(os.path.expanduser("~"), "AmanAssociates_Sera", "seraRawPayloadDump.txt"),
        os.path.join(os.path.dirname(__file__), "..", "seraRawPayloadDump.txt"),
        os.path.join(os.path.dirname(__file__), "raw_dump.txt"),
        r"C:\Users\Nex\AmanAssociates_Sera\seraRawPayloadDump.txt",
        r"I:\Project Sera\APP\FST_Classifier_1\raw_dump.txt",
    ]
    for c in candidates:
        if os.path.exists(c):
            return os.path.abspath(c)
    return path


def load_master_db_client_names():
    """Extracts PAN -> Client Name mappings from master.db if available."""
    pan_names = {}
    try:
        candidates = [
            os.path.join(os.path.expanduser("~"), "AmanAssociates_Sera", "master.db"),
            os.path.join("..", "master.db"),
            os.path.join(os.path.dirname(__file__), "..", "master.db"),
            r"C:\Users\Nex\AmanAssociates_Sera\master.db",
            r"c:\Users\Nex\Downloads\Project Sera\APP\master.db",
            r"I:\Project Sera\APP\master.db"
        ]
        salt_candidates = [
            os.path.join(os.path.expanduser("~"), "AmanAssociates_Sera", "sera.salt"),
            os.path.join("..", "sera.salt"),
            os.path.join(os.path.dirname(__file__), "..", "sera.salt"),
            r"C:\Users\Nex\AmanAssociates_Sera\sera.salt",
            r"c:\Users\Nex\Downloads\Project Sera\APP\sera.salt",
        ]
        key_candidates = [
            os.path.join(os.path.expanduser("~"), "AmanAssociates_Sera", "sera.key"),
            os.path.join("..", "sera.key"),
            os.path.join(os.path.dirname(__file__), "..", "sera.key"),
            r"C:\Users\Nex\AmanAssociates_Sera\sera.key",
            r"c:\Users\Nex\Downloads\Project Sera\APP\sera.key",
        ]

        app_dirs = [
            os.path.abspath(os.path.join(os.path.dirname(__file__), "..")),
            r"c:\Users\Nex\Downloads\Project Sera\APP",
            r"I:\Project Sera\APP"
        ]
        for ad in app_dirs:
            if os.path.exists(ad) and ad not in sys.path:
                sys.path.insert(0, ad)

        import security
        from database import SeraDatabase

        for mdb in candidates:
            if os.path.exists(mdb):
                k_file = next((k for k in key_candidates if os.path.exists(k)), None)
                s_file = next((s for s in salt_candidates if os.path.exists(s)), None)
                if k_file and s_file:
                    with open(k_file, 'r', encoding='utf-8') as f:
                        pwd = f.read().strip()
                    salt = security.load_salt(s_file)
                    hex_key = security.derive_key_hex(pwd, salt)
                    db = SeraDatabase(mdb, hex_key, defer_startup_maintenance=True)
                    clients = db.search_clients("")
                    mcl_cols = db.get_mcl_columns()
                    id_col_ids = [c["id"] for c in mcl_cols if c.get("is_identity")]
                    for c in clients:
                        vals = c.get("values", {})
                        names = [str(vals.get(cid, "")).strip() for cid in id_col_ids if vals.get(cid)]
                        c_name = " ".join(names).strip()
                        if not c_name or c_name == "Client Profile":
                            c_name = c.get("name") or ""
                        for val in vals.values():
                            if isinstance(val, str) and re.match(r'^[A-Z]{5}[0-9]{4}[A-Z]$', val.strip().upper()):
                                pan_names[val.strip().upper()] = c_name
                    break
    except Exception:
        pass
    return pan_names


GLOBAL_DB_NAMES = load_master_db_client_names()


def parse_dump(filepath):
    filepath = resolve_target_dump(filepath)
    if not os.path.exists(filepath):
        return []
        
    entries = []
    current_entry = None
    json_lines = []
    in_json = False

    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            stripped = line.strip()
            if stripped.startswith("CAPTURE DUMP ENTRY #"):
                if current_entry:
                    if json_lines:
                        try: current_entry['json'] = json.loads("".join(json_lines))
                        except: current_entry['json'] = {"error": "Failed to parse JSON"}
                    entries.append(current_entry)
                entry_num = stripped.split("#")[1].strip()
                current_entry = {"Entry #": entry_num}
                json_lines = []
                in_json = False
                continue

            if not current_entry: continue
            if stripped == "RAW JSON PAYLOAD:":
                in_json = True
                continue
            if stripped.startswith("===") or stripped.startswith("---"):
                continue

            if in_json:
                json_lines.append(line)
            else:
                if ":" in line:
                    key, val = line.split(":", 1)
                    current_entry[key.strip()] = val.strip()

    if current_entry:
        if json_lines:
            try: current_entry['json'] = json.loads("".join(json_lines))
            except: current_entry['json'] = {"error": "Failed to parse JSON"}
        entries.append(current_entry)

    return entries


def extract_identifiers(e):
    js = e.get("json", {}) if isinstance(e.get("json"), dict) else {}
    rp = js.get("raw_payload", {}) if isinstance(js.get("raw_payload"), dict) else {}
    
    pan = e.get("PAN")
    if not pan or pan == "N/A": pan = js.get("pan")
    if not pan: pan = rp.get("panNumber")
    if not pan: pan = rp.get("entityNum")
    if not pan: pan = rp.get("pan")
    if not pan: pan = rp.get("gstin")
    if not pan and isinstance(rp.get("data"), dict): pan = rp.get("data", {}).get("gstin")
    
    # Deep check for ITR forms JSON
    if not pan and isinstance(rp.get("ITR"), dict):
        for form_k in ["ITR1", "ITR2", "ITR3", "ITR4", "ITR5", "ITR6", "ITR7"]:
            if form_k in rp["ITR"]:
                p_info = rp["ITR"][form_k].get("PersonalInfo", {})
                if "PAN" in p_info:
                    pan = p_info["PAN"]
                    break
    
    if pan and isinstance(pan, str) and len(pan.strip()) >= 10:
        pan = pan.strip().upper()
    else:
        pan = None
        
    name = None
    # 1. Direct fields in raw_payload
    if rp.get("fullName"):
        name = rp.get("fullName")
    elif rp.get("firstName"):
        n = f"{rp.get('firstName', '')} {rp.get('midName', '')} {rp.get('lastName', '')}".strip()
        name = re.sub(r'\s+', ' ', n).replace('None', '').strip()
    elif rp.get("assesseeName"):
        name = rp.get("assesseeName")
    elif rp.get("taxPayerName"):
        name = rp.get("taxPayerName")
    elif rp.get("nameAsPerBank"):
        name = rp.get("nameAsPerBank")
    elif rp.get("bn"):
        name = rp.get("bn")
    elif isinstance(rp.get("data"), dict) and rp.get("data", {}).get("bn"):
        name = rp.get("data", {}).get("bn")
    elif isinstance(rp.get("data"), dict) and rp.get("data", {}).get("auth_name"):
        name = rp.get("data", {}).get("auth_name")
    elif rp.get("entityName"):
        name = rp.get("entityName")
    elif js.get("client_name"):
        name = js.get("client_name")
        
    # 2. Deep Assessee Name extraction from downloaded ITR files
    if not name and isinstance(rp.get("ITR"), dict):
        for form_k in ["ITR1", "ITR2", "ITR3", "ITR4", "ITR5", "ITR6", "ITR7"]:
            if form_k in rp["ITR"]:
                p_info = rp["ITR"][form_k].get("PersonalInfo", {})
                a_name = p_info.get("AssesseeName", {})
                if isinstance(a_name, dict):
                    name_parts = [a_name.get("FirstName", ""), a_name.get("MiddleName", ""), a_name.get("SurNameOrOrgName", "")]
                    name = " ".join([p for p in name_parts if p]).strip()
                v_info = rp["ITR"][form_k].get("Verification", {}).get("Declaration", {})
                if not name and "AssesseeVerName" in v_info:
                    name = v_info.get("AssesseeVerName")
                break
                
    # 3. Header client ID extraction
    if not name and e.get("Client ID"):
        c_str = e.get("Client ID", "")
        m = re.search(r'\(([^)]+)\)', c_str)
        if m:
            cand = m.group(1).replace("//", "").strip()
            if cand and not re.match(r'^[A-Z]{5}[0-9]{4}[A-Z]$', cand):
                name = cand
                
    # 4. Master DB lookup
    if not name and pan and pan in GLOBAL_DB_NAMES:
        name = GLOBAL_DB_NAMES[pan]
        
    ack = None
    # 15-digit Government Ack prioritization
    for k in ["ackNumber", "arnNumber", "arn", "ackNum", "transactionNo"]:
        val = rp.get(k) or js.get(k)
        if val and isinstance(val, (str, int)):
            s_val = str(val).strip()
            if re.match(r'^\d{15}$', s_val):
                ack = s_val
                break
    if not ack:
        ack = rp.get("arnNumber") or rp.get("ackNum") or js.get("arn")
    if not ack:
        hdr_ack = e.get("ARN / Ack No", "")
        if hdr_ack and not hdr_ack.startswith("PROFILE") and "EVERIFY" not in hdr_ack and "ITR-" not in hdr_ack:
            ack = hdr_ack
            
    return pan, name, ack


def resolve_identities_without_client_id(entries):
    entities = {} 
    ack_to_pan = {}
    
    # First pass: collect PANs and names
    for e in entries:
        pan, name, ack = extract_identifiers(e)
        if pan:
            if pan not in entities: entities[pan] = {"names": set(), "entries": []}
            if name: entities[pan]["names"].add(name)
            if ack: ack_to_pan[ack] = pan
                
    current_temporal_pan = None
    
    # Second pass: bind temporal entries
    for e in entries:
        pan, name, ack = extract_identifiers(e)
        if not pan:
            if ack and ack in ack_to_pan: pan = ack_to_pan[ack]
            else: pan = current_temporal_pan
        else:
            current_temporal_pan = pan 
            
        e["assigned_pan"] = pan if pan else "UNKNOWN"
        e["assigned_ack"] = ack if ack else "N/A"
        
        target_pan = e["assigned_pan"]
        if target_pan not in entities:
            entities[target_pan] = {"names": set(), "entries": []}
            
        entities[target_pan]["entries"].append(e)
        if name: entities[target_pan]["names"].add(name)
        elif target_pan in GLOBAL_DB_NAMES:
            entities[target_pan]["names"].add(GLOBAL_DB_NAMES[target_pan])
                
    return entities


def classify_bank_entry(rp):
    bank_name = rp.get("bankName", "BANK ACCOUNT")
    acc_validity = rp.get("accValidity", "")
    acc_status = rp.get("accountStatus", "")
    active_flag = rp.get("activeFlag", "")
    refund_flag = rp.get("refundFlag", "")
    remarks = rp.get("remarks", "")
    error_cd = rp.get("errorCd", "")
    
    # 1. Validated and Open for Refund
    if acc_validity == "V" and (acc_status == "Account Valid and Open" or not acc_status):
        refund_status = "Nominated for Refund" if refund_flag == "Y" else "Validation Active (Refund Not Selected)"
        return {
            "label": f"5. Bank Status: Validated ({bank_name})",
            "color": COLORS["bank_valid"],
            "details": f"Status: Valid & Open | {refund_status} | EVC Eligible: {rp.get('evcFlag', 'N')}"
        }
        
    # 2. Validated with Restriction / Warning
    if acc_validity == "V" and ("Invalid" in acc_status or "NAME_MATCH" in remarks or "mismatch" in str(error_cd).lower()):
        return {
            "label": f"5. Bank Status: Validated with Warning ({bank_name})",
            "color": COLORS["bank_warn"],
            "details": f"Status: {acc_status} | Warning: {remarks or str(error_cd)[:60]} | Refund Cap: <50L>"
        }
        
    # 3. Disabled / Legacy Inactive Account
    if active_flag == "D" or (acc_validity == "I" and "Linkage failed" in remarks and not error_cd):
        return {
            "label": f"5. Bank Status: Inactive / Legacy Account ({bank_name})",
            "color": COLORS["bank_disabled"],
            "details": f"Status: Disabled/Historical Account | Merged/Closed Bank Record | ActiveFlag: D"
        }
        
    # 4. Revalidation Required / Rejection
    fail_reason = remarks or error_cd or acc_status or "Validation Failed"
    return {
        "label": f"5. Bank Status: Revalidation Required ({bank_name})",
        "color": COLORS["bank_failed"],
        "details": f"Status: {acc_status or 'Validation Inactive'} | Reason: {str(fail_reason)[:75]}"
    }


def analyze_lifecycle(entities):
    summary_list = [] 
    for pan, data in entities.items():
        if pan == "UNKNOWN": continue 
        
        # Deduplicate and format client names cleanly
        raw_names = [n.strip() for n in data["names"] if n and str(n).strip()]
        unique_names = []
        for n in raw_names:
            if not any(n.lower() == un.lower() for un in unique_names):
                unique_names.append(n)
        names = "\n".join(unique_names) if unique_names else (GLOBAL_DB_NAMES.get(pan, "") or "N/A")
        
        has_submit, has_gst, has_itr_everified = False, False, False
        
        for e in data["entries"]:
            rp = e.get("json", {}).get("raw_payload", {})
            url = e.get("json", {}).get("url", "")
            
            # 4. GST Submit
            if "formdetails" in url and rp.get("data", {}).get("status") == "FIL":
                has_gst = True
                summary_list.append({
                    "cat": "4. GST Return Filed & E-Verified", "color": COLORS["cat4"],
                    "entries": f"Entry #{e['Entry #']}", "pan": pan, "name": names,
                    "ack": e.get("assigned_ack"), "details": "GSTR-1 successfully filed & authenticated via EVC."
                })
                
            # 5. Bank Account Status & Pre-Validation Analysis
            if "auth/getEntity" in url and ("bankName" in rp or "accValidity" in rp):
                bank_info = classify_bank_entry(rp)
                summary_list.append({
                    "cat": bank_info["label"], "color": bank_info["color"],
                    "entries": f"Entry #{e['Entry #']}", "pan": pan, "name": names,
                    "ack": "N/A", "details": bank_info["details"]
                })
                    
        submit_events = [e for e in data["entries"] if "submit/wzrd" in e.get("json",{}).get("url","") and e.get("json",{}).get("raw_payload",{}).get("httpStatus") == "ACCEPTED"]
        everify_events = [e for e in data["entries"] if "validateOTP" in e.get("json",{}).get("url","") and (e.get("json",{}).get("raw_payload",{}).get("status") == "SUCCESS" or e.get("json",{}).get("raw_payload",{}).get("code") == "OTP VALIDATED")]
        
        handled_everify_acks = set()
        
        # 1 & 2. Handle ITR Submits
        if submit_events:
            has_submit = True
            for sub_e in submit_events:
                sub_ack = sub_e.get("assigned_ack")
                
                # Check for matching e-verification by ACK or by SAME CLIENT SESSION
                matching_ev = next((ev for ev in everify_events if (ev.get("assigned_ack") == sub_ack or ev.get("assigned_pan") == pan)), None)
                
                if matching_ev:
                    has_itr_everified = True
                    handled_everify_acks.add(matching_ev.get("assigned_ack"))
                    handled_everify_acks.add(sub_ack)
                    ev_txn = matching_ev.get("json",{}).get("raw_payload",{}).get("aadhaarTxnId") or matching_ev.get('ARN / Ack No') or sub_ack
                    summary_list.append({ 
                        "cat": "2. File Submitted & E-Verified (ITR)", "color": COLORS["cat2"],
                        "entries": f"Submit #{sub_e['Entry #']}, E-Verify #{matching_ev['Entry #']}",
                        "pan": pan, "name": names, "ack": sub_ack,
                        "details": f"Complete ITR lifecycle: JSON submission + OTP E-Verification via Aadhaar (Txn: {ev_txn})"
                    })
                else:
                    summary_list.append({ 
                        "cat": "1. File Submitted (NOT E-Verified)", "color": COLORS["cat1"],
                        "entries": f"Entry #{sub_e['Entry #']}", "pan": pan, "name": names, 
                        "ack": f"{sub_ack} (Txn: {sub_e.get('ARN / Ack No')})",
                        "details": "ITR Return submitted via wizard with HTTP status ACCEPTED, but EVC/OTP e-verification is pending."
                    })
                    
        # Handle Standalone ITR E-Verification (when return was submitted in prior session)
        for ev in everify_events:
            ev_rp = ev.get("json", {}).get("raw_payload", {})
            ev_ack = ev.get("assigned_ack")
            if ev_ack not in handled_everify_acks:
                has_itr_everified = True
                form_num = ev_rp.get("formCd", "ITR")
                summary_list.append({ 
                    "cat": "2. File Submitted & E-Verified (ITR)", "color": COLORS["cat2"],
                    "entries": f"E-Verify #{ev['Entry #']} (AY {ev_rp.get('assessmntYr', '')})",
                    "pan": pan, "name": names, "ack": ev_ack,
                    "details": f"ITR-{form_num} Return successfully e-verified via Aadhaar OTP (Txn: {ev.get('ARN / Ack No')})"
                })
                    
        # 3. Check for standalone Bank Verification
        for ev in everify_events:
            if ev.get("json",{}).get("raw_payload",{}).get("moduleCode") == "NON-ITR":
                if not has_submit and not has_itr_everified: 
                    summary_list.append({ 
                        "cat": "3. Bank Account E-Verified (NO Return Submitted)", "color": COLORS["cat3"],
                        "entries": f"Entry #{ev['Entry #']}", "pan": pan, "name": names,
                        "ack": ev.get("assigned_ack"), "details": "Bank Account Re-validation e-verified independently via Aadhaar OTP."
                    })
                    
        # 6. Category 6: Visited but no return submission (Only if NO submit and NO everify)
        if not has_submit and not has_gst and not has_itr_everified:
            visit_entries = [f"#{e['Entry #']}" for e in data["entries"]]
            summary_list.append({
                "cat": "6. Visited But No Return Submission", "color": COLORS["cat6"],
                "entries": ", ".join(visit_entries), "pan": pan, "name": names, "ack": "N/A",
                "details": f"Client had {len(data['entries'])} interactions (e.g., profile sync, wizard prep, bank checks) but ZERO ITR or GST final submissions."
            })
            
        # 7. Category 7: All visits enumerated
        visit_details = []
        for e in data["entries"]:
            url = e.get("json", {}).get("url", "Unknown Endpoint")
            ts = e.get("Timestamp", "")
            visit_details.append(f"[#{e['Entry #']} @ {ts[-14:-6]}] - {url.split('/')[-1]}")
            
        summary_list.append({
            "cat": "7. Visited Site (All Visits Enumerated)", "color": COLORS["cat7"],
            "entries": f"Total: {len(data['entries'])}", "pan": pan, "name": names, "ack": "N/A",
            "details": "\n".join(visit_details)
        })

    summary_list.sort(key=lambda x: x["cat"])
    return summary_list


def create_excel(entries, entities, summary_list, out_path):
    wb = openpyxl.Workbook()
    
    ws_sum = wb.active
    ws_sum.title = "Action Summary"
    ws_sum.views.sheetView[0].showGridLines = True
    
    ws_sum["A1"] = "PROJECT SERA — AUTOMATED FST PAYLOAD CLASSIFICATION"
    ws_sum["A1"].font = TITLE_FONT
    ws_sum["A2"] = f"Generated at {time.strftime('%Y-%m-%d %H:%M:%S')} - Live Tracker Active ({len(entries)} Entries Parsed)"
    ws_sum["A2"].font = SUB_TITLE_FONT
    
    headers = ["Lifecycle Category", "Entries Involved", "PAN / GSTIN", "Client Name / Entity", "ARN / Ack Number", "Lifecycle Verification Details"]
    ws_sum.row_dimensions[4].height = 26
    for col_idx, h in enumerate(headers, 1):
        c = ws_sum.cell(row=4, column=col_idx, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_idx = 5
    for row_data in summary_list:
        ws_sum.row_dimensions[row_idx].height = 90 if "7. Visited" in row_data["cat"] else 40
        
        c1 = ws_sum.cell(row=row_idx, column=1, value=row_data["cat"])
        c2 = ws_sum.cell(row=row_idx, column=2, value=row_data["entries"])
        c3 = ws_sum.cell(row=row_idx, column=3, value=row_data["pan"])
        c4 = ws_sum.cell(row=row_idx, column=4, value=row_data["name"])
        c5 = ws_sum.cell(row=row_idx, column=5, value=row_data["ack"])
        c6 = ws_sum.cell(row=row_idx, column=6, value=row_data["details"])
        
        for c in [c1, c2, c3, c4, c5, c6]:
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.font = Font(name="Calibri", size=10)
            
        c1.fill = row_data["color"]
        c1.font = Font(name="Calibri", size=10, bold=True)
        
        row_idx += 1

    ws_sum.column_dimensions['A'].width = 46
    ws_sum.column_dimensions['B'].width = 22
    ws_sum.column_dimensions['C'].width = 18
    ws_sum.column_dimensions['D'].width = 32
    ws_sum.column_dimensions['E'].width = 34
    ws_sum.column_dimensions['F'].width = 80

    ws_all = wb.create_sheet(title="All Logged Entries")
    all_headers = ["Entry #", "Timestamp (UTC)", "Assigned PAN", "Assigned Ack", "Endpoint URL"]
    ws_all.row_dimensions[1].height = 20
    for col_idx, h in enumerate(all_headers, 1):
        c = ws_all.cell(row=1, column=col_idx, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    r_idx = 2
    for e in entries:
        c1 = ws_all.cell(row=r_idx, column=1, value=e.get("Entry #"))
        c2 = ws_all.cell(row=r_idx, column=2, value=e.get("Timestamp"))
        c3 = ws_all.cell(row=r_idx, column=3, value=e.get("assigned_pan"))
        c4 = ws_all.cell(row=r_idx, column=4, value=e.get("assigned_ack"))
        c5 = ws_all.cell(row=r_idx, column=5, value=e.get("json", {}).get("url", ""))
        
        for c in [c1, c2, c3, c4, c5]:
            c.border = THIN_BORDER
            c.font = Font(name="Calibri", size=9)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            
        c1.alignment = Alignment(horizontal="center", vertical="center")
        r_idx += 1

    ws_all.column_dimensions['B'].width = 24
    ws_all.column_dimensions['C'].width = 18
    ws_all.column_dimensions['D'].width = 25
    ws_all.column_dimensions['E'].width = 75

    # Safe Save with Retry & Fallback if file is open in Excel
    saved = False
    max_retries = 3
    for attempt in range(max_retries):
        try:
            wb.save(out_path)
            wb.close()
            saved = True
            break
        except PermissionError:
            time.sleep(0.4)
        except Exception as e:
            print(f"[-] Error saving {out_path}: {e}")
            break

    if not saved:
        # Save to fallback mirror file
        dir_name = os.path.dirname(out_path) or "."
        base_name = os.path.basename(out_path)
        name_no_ext, ext = os.path.splitext(base_name)
        fallback_path = os.path.join(dir_name, f"{name_no_ext}_latest{ext}")
        try:
            wb.save(fallback_path)
            wb.close()
            print(f"[!] '{out_path}' is currently open in Excel.")
            print(f"[+] Saved updated copy to '{fallback_path}'. Will sync to '{out_path}' once Excel is closed.")
            return True, fallback_path
        except Exception as e:
            print(f"[-] Fallback save failed: {e}")
            return False, None

    return True, out_path


def process_data(input_dump, output_excel):
    try:
        real_dump_path = resolve_target_dump(input_dump)
        entries = parse_dump(real_dump_path)
        if not entries: return False, None
        entities = resolve_identities_without_client_id(entries)
        summary = analyze_lifecycle(entities)
        return create_excel(entries, entities, summary, output_excel)
    except Exception as e:
        print(f"[-] Error processing dump: {e}")
        return False, None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("input_dump", nargs="?", default="..\\seraRawPayloadDump.txt", help="Path to raw dump text file")
    parser.add_argument("output_excel", nargs="?", default="payload_report.xlsx", help="Path to output Excel tracker")
    parser.add_argument("--watch", action="store_true", help="Monitor the dump file for changes and auto-update")
    args = parser.parse_args()

    target_dump = resolve_target_dump(args.input_dump)

    if args.watch:
        print(f"[*] Live Tracker Mode Activated.")
        print(f"[*] Watching '{target_dump}' for updates...")
        
        last_mtime = -1
        pending_retry = False
        while True:
            try:
                if os.path.exists(target_dump):
                    current_mtime = os.path.getmtime(target_dump)
                    if current_mtime > last_mtime or pending_retry:
                        if last_mtime != -1 and not pending_retry:
                            print(f"\n[+] Detected new payloads at {time.strftime('%H:%M:%S')}. Recompiling tracker...")
                        elif last_mtime == -1:
                            print(f"[*] Compiling initial tracker dump...")
                            
                        success, actual_path = process_data(target_dump, args.output_excel)
                        if success:
                            if actual_path == args.output_excel:
                                print(f"[+] Tracker successfully updated: {args.output_excel}")
                                last_mtime = current_mtime
                                pending_retry = False
                            else:
                                # Saved to fallback mirror while main file was open in Excel
                                pending_retry = True
                        else:
                            pending_retry = True
                else:
                    if last_mtime != -2:
                        print(f"[-] Waiting for '{target_dump}' to be created...")
                        last_mtime = -2
            except KeyboardInterrupt:
                print("\n[*] Live Tracker Stopped.")
                sys.exit(0)
            except Exception as e:
                print(f"[-] Watcher error: {e}")
                
            time.sleep(2)
    else:
        print(f"[*] Processing dump from '{target_dump}'...")
        success, actual_path = process_data(target_dump, args.output_excel)
        if success: print(f"[+] Success! Tracker saved to {actual_path}")


if __name__ == "__main__":
    main()
