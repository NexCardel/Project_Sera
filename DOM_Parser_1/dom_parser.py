import os
import sys
import re
import json
import time
import argparse
try:
    import sqlcipher3.dbapi2 as sqlite3
except ImportError:
    import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- STYLING CONSTANTS (Material Palette) ---
HEADER_FILL = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid") # Deep Navy
HEADER_FONT = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Segoe UI", size=15, bold=True, color="1A365D")
SUB_TITLE_FONT = Font(name="Segoe UI", size=10, italic=True, color="4A5568")
SECTION_FONT = Font(name="Segoe UI", size=11, bold=True, color="2D3748")

THIN_BORDER = Border(
    left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0')
)

COLORS = {
    "cat1": PatternFill(start_color="DEF7EC", end_color="DEF7EC", fill_type="solid"), # 1. Filed & Verified (Green)
    "cat2": PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid"), # 2. Submitted Pending (Yellow)
    "cat3": PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid"), # 3. SDC Timelines (Cyan/Blue)
    "cat4": PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid"), # 4. Drafts & Schedules (Indigo)
    "cat5": PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid"), # 5. Taxpayer Identity (Purple)
    "cat6": PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid"), # 6. Navigation Journey (Slate)
    "status_completed": PatternFill(start_color="DEF7EC", end_color="DEF7EC", fill_type="solid"),
    "status_abrupt": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    "status_active": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
}

def get_db_hex_key():
    """Derives SQLCipher key from sera.key / default password and sera.salt."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    app_dir = os.path.abspath(os.path.join(base_dir, ".."))
    
    salt_candidates = [
        os.path.join(app_dir, "sera.salt"),
        os.path.join("..", "sera.salt"),
        os.path.join(os.path.expanduser("~"), "AmanAssociates_Sera", "sera.salt"),
        os.path.join(app_dir, "salt.bin"),
        os.path.join(os.path.expanduser("~"), "AmanAssociates_Sera", "salt.bin"),
    ]
    key_candidates = [
        os.path.join(app_dir, "sera.key"),
        os.path.join("..", "sera.key"),
        os.path.join(os.path.expanduser("~"), "AmanAssociates_Sera", "sera.key"),
    ]
    
    salt = None
    for s_path in salt_candidates:
        if os.path.exists(s_path):
            try:
                with open(s_path, "rb") as f:
                    salt = f.read()
                if salt: break
            except Exception: pass
            
    pwd = "admin123"
    for k_path in key_candidates:
        if os.path.exists(k_path):
            try:
                with open(k_path, "r", encoding="utf-8") as f:
                    p = f.read().strip()
                if p:
                    pwd = p
                    break
            except Exception: pass
            
    if salt:
        try:
            if app_dir not in sys.path:
                sys.path.insert(0, app_dir)
            import security
            return security.derive_key_hex(pwd, salt)
        except Exception:
            pass
    return None

def resolve_target_dump(input_arg=None):
    """Finds the most appropriate dump source file or rawPayload.db."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    app_dir = os.path.abspath(os.path.join(base_dir, ".."))
    live_app_dir = os.path.join(os.path.expanduser("~"), "AmanAssociates_Sera")
    
    if input_arg and os.path.exists(input_arg):
        return os.path.abspath(input_arg)

    candidates = [
        os.path.join(live_app_dir, "rawPayload.db"),
        os.path.join(app_dir, "rawPayload.db"),
        os.path.join(live_app_dir, "seraRawPayloadDump.txt"),
        os.path.join(app_dir, "seraRawPayloadDump.txt"),
        os.path.join(app_dir, "Raw_Payload_Dump", "seraRawPayloadDumpBackup.txt")
    ]

    for c in candidates:
        if os.path.exists(c):
            return os.path.abspath(c)

    return os.path.join(app_dir, "rawPayload.db")

def parse_entries_from_sqlite(db_path):
    """Fetches and parses raw DOM captures and SDC session timelines directly from rawPayload.db."""
    if not os.path.exists(db_path):
        return [], []
    entries = []
    timelines = []
    try:
        conn = sqlite3.connect(db_path)
        hex_key = get_db_hex_key()
        if hex_key:
            conn.execute(f"PRAGMA key = \"x'{hex_key}'\";")
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        # 1. Fetch tracker_dump entries
        try:
            cur.execute("""
                SELECT id, client_id, unassigned_identity, portal, period_label, 
                       arn_number, capture_method, status, raw_payload_json, created_at
                FROM tracker_dump
                ORDER BY id ASC
            """)
            rows = cur.fetchall()
            for r in rows:
                entry_id = str(r["id"])
                raw_json = {}
                if r["raw_payload_json"]:
                    try:
                        raw_json = json.loads(r["raw_payload_json"])
                    except Exception:
                        raw_json = {"raw": r["raw_payload_json"]}
                
                entry = {
                    "Entry #": entry_id,
                    "Timestamp": r["created_at"] or "",
                    "Portal": r["portal"] or "Portal",
                    "PAN": r["unassigned_identity"] or "",
                    "Client ID": str(r["client_id"] or ""),
                    "ARN / Ack No": r["arn_number"] or "N/A",
                    "Period": r["period_label"] or "",
                    "Method": r["capture_method"] or "DOM_Tracker",
                    "Status": r["status"] or "captured",
                    "json": raw_json
                }
                entries.append(entry)
        except Exception as e:
            print(f"[DOM_Parser_1] tracker_dump read notice: {e}")

        # 2. Fetch sdc_session_timelines
        try:
            cur.execute("""
                SELECT session_id, client_id, pan, client_name, portal, status,
                       start_time, end_time, total_steps, timeline_json, last_updated
                FROM sdc_session_timelines
                ORDER BY last_updated DESC
            """)
            s_rows = cur.fetchall()
            for sr in s_rows:
                tl_data = []
                if sr["timeline_json"]:
                    try:
                        tl_data = json.loads(sr["timeline_json"])
                    except Exception:
                        pass
                timelines.append({
                    "session_id": sr["session_id"],
                    "client_id": sr["client_id"],
                    "pan": sr["pan"] or "",
                    "client_name": sr["client_name"] or "",
                    "portal": sr["portal"] or "income tax",
                    "status": sr["status"] or "completed",
                    "start_time": sr["start_time"] or "",
                    "end_time": sr["end_time"] or "",
                    "total_steps": sr["total_steps"] or len(tl_data),
                    "timeline": tl_data,
                    "created_at": sr["last_updated"] or ""
                })
        except Exception as e:
            print(f"[DOM_Parser_1] sdc_session_timelines read notice: {e}")

        conn.close()
    except Exception as e:
        print(f"[DOM_Parser_1] SQLite connection notice: {e}")
    return entries, timelines

def parse_entries_from_text(filepath):
    """Parses entries from seraRawPayloadDump.txt."""
    if not os.path.exists(filepath):
        return [], []
    entries = []
    current_entry = None
    json_lines = []
    in_json = False

    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            stripped = line.strip()
            if stripped.startswith("CAPTURE DUMP ENTRY #"):
                if current_entry:
                    if json_lines:
                        try: current_entry['json'] = json.loads("".join(json_lines))
                        except: current_entry['json'] = {"error": "Failed to parse JSON"}
                    entries.append(current_entry)
                entry_num = stripped.split("#")[1].strip()
                current_entry = {"Entry #": entry_num}
                json_lines = []
                in_json = False
                continue

            if not current_entry: continue
            if stripped == "RAW JSON PAYLOAD:":
                in_json = True
                continue
            if stripped.startswith("===") or stripped.startswith("---"):
                continue

            if in_json:
                json_lines.append(line)
            else:
                if ":" in line:
                    key, val = line.split(":", 1)
                    current_entry[key.strip()] = val.strip()

    if current_entry:
        if json_lines:
            try: current_entry['json'] = json.loads("".join(json_lines))
            except: current_entry['json'] = {"error": "Failed to parse JSON"}
        entries.append(current_entry)

    return entries, []

def load_data(filepath=None):
    """Loads entries and timelines from SQLite (preferred) or text dump."""
    target = resolve_target_dump(filepath)
    if not target or not os.path.exists(target):
        return [], []
    if target.endswith(".db"):
        return parse_entries_from_sqlite(target)
    else:
        return parse_entries_from_text(target)

def extract_dom_details(entry):
    """
    Intelligently extracts and normalizes all visual DOM details from scraped_data
    and raw_payload.
    """
    raw = entry.get("json") or {}
    rp = raw.get("raw_payload") if isinstance(raw.get("raw_payload"), dict) else raw
    scraped = raw.get("scraped_data") or rp.get("scraped_data") or {}

    # 1. PAN and GSTIN
    pan = entry.get("PAN") or raw.get("pan") or rp.get("pan") or ""
    gstin = ""
    summary = scraped.get("summary_labels") or {}
    if "GSTIN" in summary:
        gstin = summary["GSTIN"]
        if not pan and len(gstin) == 15:
            pan = gstin[2:12]

    for nb in scraped.get("ng_binds") or []:
        if "gstin" in str(nb.get("expr", "")).lower() and nb.get("value"):
            gstin = str(nb["value"]).strip().upper()
            if not pan and len(gstin) == 15:
                pan = gstin[2:12]
            break

    # 2. Assembled Taxpayer Name
    form_fields = scraped.get("form_fields") or {}
    f_name, m_name, l_name = "", "", ""
    for k, v in form_fields.items():
        k_lower = k.lower()
        val = str(v).strip()
        if not val or len(val) > 60: continue
        if "firstname" in k_lower or "first_name" in k_lower:
            f_name = val
        elif "middlename" in k_lower or "middle_name" in k_lower:
            m_name = val
        elif "surname" in k_lower or "lastname" in k_lower or "last_name" in k_lower or "orgname" in k_lower:
            l_name = val

    assembled_form_name = " ".join(part for part in [f_name, m_name, l_name] if part).strip() if (f_name or l_name) else ""

    client_name = (
        assembled_form_name 
        or summary.get("Legal Name") 
        or summary.get("Trade Name")
        or raw.get("client_name") 
        or raw.get("name") 
        or raw.get("taxpayer_name")
        or rp.get("client_name")
        or ""
    )

    # Filter section headings and UI noise names
    NOISE_NAMES = {"menu", "close", "help", "refresh", "download", "button", "profile", "login", "logout", "submit", "first name", "last name", "part a", "general information", "a1. first name"}
    if client_name and (re.match(r"^(?:[A-Z]?\d{1,3}\.|\d+[\.\s-])", client_name) or client_name.strip().lower() in NOISE_NAMES):
        client_name = assembled_form_name or ""

    # 3. Form / Filing Type
    filing_type = (
        raw.get("filing_type") 
        or rp.get("filing_type") 
        or summary.get("Form") 
        or ""
    )
    if not filing_type or filing_type == "Filing Confirmation" or filing_type == "Dashboard / Profile":
        bc_text = str(scraped.get("breadcrumbs") or raw.get("dom_breadcrumbs") or "")
        itr_m = re.search(r"\b(ITR-[1-7][A-Z]?|GSTR-1\/IFF|GSTR-[1234789][AB]?|CMP-08|GSTR-9|GSTR-9C|Form\s*16[A]?|Form\s*24Q|Form\s*26Q)\b", bc_text, re.I)
        if itr_m:
            filing_type = itr_m.group(1).upper()
        else:
            url_text = str(raw.get("url") or rp.get("url") or "")
            gst_url_m = re.search(r"(?:returns\/auth\/gstr|returns\/auth\/|gstr[-_]?)(1|3b|4|9|9c|cmp08)", url_text, re.I)
            if gst_url_m:
                g_code = gst_url_m.group(1).upper()
                filing_type = "CMP-08" if g_code == "CMP08" else f"GSTR-{g_code}"
            else:
                url_m = re.search(r"fo-itr([1-7][a-z]?)", url_text, re.I)
                if url_m:
                    filing_type = f"ITR-{url_m.group(1).upper()}"

    if not filing_type:
        filing_type = "ITR Return"

    # 4. Period / Assessment Year
    period = entry.get("Period") or raw.get("period_label") or rp.get("period") or ""
    if not period:
        fy = summary.get("FY")
        tax_p = summary.get("Tax Period") or summary.get("Return Period")
        if tax_p and fy:
            period = f"{tax_p} (FY {fy})"
        elif fy:
            period = f"FY {fy}"
        elif summary.get("Assessment Year"):
            period = f"AY {summary['Assessment Year']}"

    # 5. ARN / Ack Number
    arn = entry.get("ARN / Ack No") or raw.get("arn") or rp.get("arn") or "N/A"
    if arn == "Feedback" or arn == "wledgement" or not re.search(r"\d", str(arn)):
        arn = "N/A"

    # 6. Breadcrumbs, History & Confirmation
    breadcrumbs = scraped.get("breadcrumbs") or raw.get("dom_breadcrumbs") or ""
    link_history = raw.get("site_link_history") or rp.get("site_link_history") or ""
    conf_text = scraped.get("confirmation_text") or raw.get("confirmation_message") or rp.get("confirmation_message") or ""

    return {
        "pan": pan.strip().upper() if pan else "UNKNOWN",
        "gstin": gstin.strip().upper() if gstin else "",
        "name": client_name.strip() if client_name else "Unknown Assessee",
        "form": filing_type,
        "period": period or "AY 2026-27",
        "arn": arn,
        "status": summary.get("Status") or raw.get("status") or entry.get("Status") or "Captured",
        "breadcrumbs": breadcrumbs,
        "link_history": link_history,
        "confirmation_text": conf_text,
        "url": raw.get("url") or rp.get("url") or "",
        "timestamp": entry.get("Timestamp") or raw.get("timestamp") or "",
        "entry_num": entry.get("Entry #", "")
    }

def resolve_identities_bi_directionally(entries):
    """
    FST Classifier Algorithm:
    Runs bi-directional (forward + backward) sweep to propagate taxpayer PAN
    and legal name across all setup and navigation entries in the same session.
    """
    n = len(entries)
    if n == 0:
        return {}

    direct_pans = []
    direct_names = []
    direct_gstins = []
    session_bounds = [False] * n

    for i, e in enumerate(entries):
        details = extract_dom_details(e)
        pan = details["pan"] if details["pan"] != "UNKNOWN" else ""
        name = details["name"] if details["name"] != "Unknown Assessee" else ""
        gstin = details["gstin"]

        url_str = str(e.get("json", {}).get("url") or "").lower()
        if "logout" in url_str or "signout" in url_str or "sign-out" in url_str:
            session_bounds[i] = True

        direct_pans.append(pan)
        direct_names.append(name)
        direct_gstins.append(gstin)

    # 1. Forward Propagation
    curr_pan = ""
    curr_name = ""
    curr_gstin = ""
    fwd_pans = []
    fwd_names = []
    fwd_gstins = []

    for i in range(n):
        if session_bounds[i]:
            curr_pan = ""
            curr_name = ""
            curr_gstin = ""

        if direct_pans[i]: curr_pan = direct_pans[i]
        if direct_names[i]: curr_name = direct_names[i]
        if direct_gstins[i]: curr_gstin = direct_gstins[i]

        fwd_pans.append(curr_pan)
        fwd_names.append(curr_name)
        fwd_gstins.append(curr_gstin)

    # 2. Backward Sweep
    rev_pan = ""
    rev_name = ""
    rev_gstin = ""
    final_pans = [""] * n
    final_names = [""] * n
    final_gstins = [""] * n

    for i in range(n - 1, -1, -1):
        if session_bounds[i]:
            rev_pan = ""
            rev_name = ""
            rev_gstin = ""

        if direct_pans[i]: rev_pan = direct_pans[i]
        if direct_names[i]: rev_name = direct_names[i]
        if direct_gstins[i]: rev_gstin = direct_gstins[i]

        final_pans[i] = direct_pans[i] or fwd_pans[i] or rev_pan or "UNKNOWN"
        final_names[i] = direct_names[i] or fwd_names[i] or rev_name or "Unknown Assessee"
        final_gstins[i] = direct_gstins[i] or fwd_gstins[i] or rev_gstin or ""

    # Group into Entities
    entities = {}
    for i, e in enumerate(entries):
        p = final_pans[i]
        g = final_gstins[i]
        target_key = p if p != "UNKNOWN" else (g if g else f"UNASSIGNED_ENTRY_{i+1}")
        resolved_name = final_names[i]

        if target_key not in entities:
            entities[target_key] = {
                "pan": p,
                "gstin": g,
                "names": set(),
                "entries": []
            }
        
        entities[target_key]["entries"].append(e)
        if resolved_name != "Unknown Assessee":
            entities[target_key]["names"].add(resolved_name)
        if direct_gstins[i]:
            entities[target_key]["gstin"] = direct_gstins[i]

    return entities

def classify_entries(entries, timelines=None):
    """
    Pinpoint Multi-Layer Classifier (SDC + DOM 3-Factor Lifecycle):
    1. Bi-directional entity reconciliation.
    2. Event Pairing: Correlates Submit and e-Verify banners across taxpayer history.
    3. Categorizes into 6 distinct lifecycle groups:
       Cat 1: Filed & Verified Returns
       Cat 2: Submitted Pending e-Verification
       Cat 3: SDC Session Timelines & Clickstream Trace
       Cat 4: Return Drafts & Active Schedules
       Cat 5: Taxpayer Identity Ledger
       Cat 6: User Navigation Journey Trace
    """
    classified = {
        "cat1": [], # Filed & Verified
        "cat2": [], # Submitted (Pending)
        "cat3": [], # SDC Timelines
        "cat4": [], # Drafts & Schedules
        "cat5": {}, # Taxpayer Profiles (keyed by PAN/GSTIN)
        "cat6": []  # Navigation Journeys
    }

    # Populate Cat 3 from rawPayload.db sdc_session_timelines
    if timelines:
        for t in timelines:
            sess_id = t.get("session_id", "")
            p_name = t.get("client_name") or "Taxpayer"
            p_pan = t.get("pan") or ""
            p_portal = t.get("portal") or "income tax"
            p_status = t.get("status") or "completed"
            
            tl_steps = t.get("timeline") or []
            if not tl_steps:
                classified["cat3"].append({
                    "session_id": sess_id,
                    "client_name": p_name,
                    "pan": p_pan,
                    "portal": p_portal,
                    "session_status": p_status,
                    "step": 1,
                    "timestamp": t.get("start_time", ""),
                    "action_title": "Session Recorded",
                    "url": "Portal Session",
                    "crosshair_id": "sdc_session",
                    "parameters": f"Status: {p_status.upper()}"
                })
            else:
                for step_obj in tl_steps:
                    st_num = step_obj.get("step", 1)
                    st_time = step_obj.get("time") or step_obj.get("timestamp") or t.get("start_time", "")
                    st_title = step_obj.get("route_title") or step_obj.get("title") or "Page Navigation"
                    st_url = step_obj.get("url") or ""
                    st_crosshair = step_obj.get("crosshair") or step_obj.get("crosshair_id") or "sdc_navigation"
                    
                    # Format parameters
                    cap = step_obj.get("capture") or {}
                    param_parts = []
                    if cap.get("filing_type"): param_parts.append(f"Form: {cap['filing_type']}")
                    if cap.get("period_label"): param_parts.append(f"AY: {cap['period_label']}")
                    if cap.get("arn") and cap["arn"] != "N/A": param_parts.append(f"Ack: {cap['arn']}")
                    if cap.get("status"): param_parts.append(f"Status: {cap['status']}")
                    
                    param_str = " | ".join(param_parts) if param_parts else "-"
                    
                    classified["cat3"].append({
                        "session_id": sess_id,
                        "client_name": p_name,
                        "pan": p_pan,
                        "portal": p_portal,
                        "session_status": p_status,
                        "step": st_num,
                        "timestamp": st_time,
                        "action_title": st_title,
                        "url": st_url,
                        "crosshair_id": st_crosshair,
                        "parameters": param_str
                    })

    if not entries:
        return classified

    entities = resolve_identities_bi_directionally(entries)

    for pan_key, entity in entities.items():
        # Canonical Name Selection
        best_name = "Unknown Assessee"
        valid_names = [n for n in entity["names"] if n and n != "Unknown Assessee" and not any(t in n.lower() for t in ("menu", "close", "help", "refresh", "button", "first name", "last name"))]
        if valid_names:
            best_name = max(valid_names, key=len)

        # Populate Cat 5: Taxpayer Ledger
        forms_seen = set()
        latest_period = "AY 2026-27"
        latest_url = ""
        for e in entity["entries"]:
            d = extract_dom_details(e)
            if d["form"]: forms_seen.add(d["form"])
            if d["period"]: latest_period = d["period"]
            if d["url"]: latest_url = d["url"]

        classified["cat5"][pan_key] = {
            "pan": entity["pan"],
            "gstin": entity["gstin"],
            "name": best_name,
            "latest_period": latest_period,
            "forms_seen": forms_seen or {"ITR Return"},
            "latest_url": latest_url,
            "total_captures": len(entity["entries"])
        }

        # Multi-Event Lifecycle Pass
        submit_events = []
        everify_events = []
        draft_events = []

        for e in entity["entries"]:
            d = extract_dom_details(e)
            e_rec = {
                **e,
                **d,
                "name": best_name if best_name != "Unknown Assessee" else d["name"],
                "pan": entity["pan"] if entity["pan"] != "UNKNOWN" else d["pan"]
            }

            url_lower = (d["url"] or "").lower()
            bc_lower = (d["breadcrumbs"] or "").lower()
            conf_lower = (d.get("confirmation_text") or "").lower()
            status_lower = (d["status"] or "").lower()

            classified["cat6"].append(e_rec)

            # Signal 1: Completed & Verified Filing (Cat 1)
            is_verified = (
                ("successfully filed" in conf_lower or "verified successfully" in conf_lower or "e-verification completed" in conf_lower 
                 or "filing-success" in url_lower or "view-filed-returns" in url_lower or "fo-e-verify-now-success" in url_lower or "fo-return-success" in url_lower)
                or (status_lower in ("filed & verified", "filed and verified", "e-verified", "filed & verified (portal confirmed)", "fil"))
                or (d["arn"] != "N/A" and re.search(r"\d{7,}", str(d["arn"])) and ("filed" in status_lower or "verified" in status_lower))
                or ("gstr1/success" in url_lower or "gstr3b/summary" in url_lower)
            )

            # Signal 2: Submitted Pending e-Verification (Cat 2)
            is_submitted_pending = (
                not is_verified and (
                    ("fo-e-verify-later" in url_lower or "everifylater" in url_lower or "e-verify later" in conf_lower 
                     or "successfully submitted" in conf_lower or "download itr-v" in conf_lower or "complete-verification" in url_lower)
                    or ("submitted (pending" in status_lower or "pending e-verification" in status_lower)
                )
                and not any(nav in url_lower for nav in ("login", "dashboard", "select-status", "personal_information", "parta"))
            )

            # Signal 3: Active Schedules & Drafts (Cat 4)
            is_draft_schedule = (
                any(s in url_lower for s in ("personal_information", "parta", "schedule", "fo-itr", "foreturns", "select-status", "returns", "computation", "gross"))
                or any(s in bc_lower for s in ("personal information", "part a", "select status", "schedule", "filing returns"))
                or d.get("json", {}).get("scraped_data", {}).get("form_fields")
            )

            if is_verified:
                everify_events.append(e_rec)
            elif is_submitted_pending:
                submit_events.append(e_rec)
            elif is_draft_schedule:
                draft_events.append(e_rec)

        if everify_events:
            classified["cat1"].extend(everify_events)
        elif submit_events:
            classified["cat2"].extend(submit_events)
        
        if draft_events:
            latest_drafts_by_page = {}
            for d_rec in draft_events:
                p_key = (d_rec.get("url") or d_rec.get("form") or "").strip().split("?")[0].rstrip("/").lower()
                latest_drafts_by_page[p_key] = d_rec
            classified["cat4"].extend(latest_drafts_by_page.values())

    return classified

def generate_excel_report(classified, output_path="dom_audit_report.xlsx"):
    """
    Generates a professionally formatted multi-tab Excel audit workbook
    using openpyxl.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Count unique sessions
    sessions_seen = set()
    completed_sessions = set()
    abrupt_sessions = set()
    for row in classified.get("cat3", []):
        s_id = row.get("session_id")
        if s_id:
            sessions_seen.add(s_id)
            if row.get("session_status") == "completed":
                completed_sessions.add(s_id)
            elif row.get("session_status") == "terminated_abruptly":
                abrupt_sessions.add(s_id)

    # --- TAB 1: EXECUTIVE SUMMARY & DASHBOARD ---
    ws_summary = wb.create_sheet(title="Executive Summary")
    ws_summary.views.sheetView[0].showGridLines = True

    ws_summary["A1"] = "PROJECT SERA — SDC & DOM CLASSIFIER AUDIT REPORT"
    ws_summary["A1"].font = TITLE_FONT
    ws_summary["A2"] = f"Visual Layer DOM Tracking, SDC Timelines & Taxpayer Reconciliations on {time.strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary["A2"].font = SUB_TITLE_FONT

    ws_summary.append([]) # Row 3 blank
    ws_summary.append(["Category Metric", "Count", "Classification Description"])
    sum_hdr_row = 4
    for col in range(1, 4):
        cell = ws_summary.cell(row=sum_hdr_row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left" if col != 2 else "center", vertical="center")

    summary_rows = [
        ("1. Filed & Verified Returns", len(classified["cat1"]), "Completed & e-verified returns (confirmed via route, success banner, or Ack).", COLORS["cat1"]),
        ("2. Submitted (Pending Verification)", len(classified["cat2"]), "Submissions pending 30-day e-verification (ITR-V / e-Verify Later route).", COLORS["cat2"]),
        ("3. SDC Session Timelines & Steps", len(classified["cat3"]), f"Chronological clickstream trace ({len(sessions_seen)} unique sessions: {len(completed_sessions)} completed, {len(abrupt_sessions)} abrupt).", COLORS["cat3"]),
        ("4. Active Return Drafts & Schedules", len(classified["cat4"]), "Active computation schedules, PartA_GEN, and GSTR details.", COLORS["cat4"]),
        ("5. Reconciled Taxpayers", len(classified["cat5"]), "Unique PAN & GSTIN identity containers with resolved legal names.", COLORS["cat5"]),
        ("6. Navigation Link Journeys", len(classified["cat6"]), "Audit trail of user page-by-page link navigation timestamps.", COLORS["cat6"]),
    ]

    for idx, (cat_name, count, desc, fill_c) in enumerate(summary_rows, start=5):
        ws_summary.append([cat_name, count, desc])
        for c in range(1, 4):
            cell = ws_summary.cell(row=idx, column=c)
            cell.border = THIN_BORDER
            cell.font = Font(name="Segoe UI", size=10)
            if c == 1:
                cell.fill = fill_c
                cell.font = Font(name="Segoe UI", size=10, bold=True)
            elif c == 2:
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(name="Segoe UI", size=10, bold=True)

    # --- TAB 2: FILED & VERIFIED (CAT 1) ---
    ws_cat1 = wb.create_sheet(title="1. Filed & Verified")
    ws_cat1.views.sheetView[0].showGridLines = True
    headers_cat1 = ["Entry #", "Timestamp", "Portal", "Assessee Name", "PAN", "GSTIN", "Form", "Period", "ACK / ARN Number", "Status"]
    ws_cat1.append(headers_cat1)
    for col in range(1, len(headers_cat1) + 1):
        cell = ws_cat1.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, item in enumerate(classified["cat1"], start=2):
        ws_cat1.append([
            item.get("entry_num", ""), item.get("timestamp", ""), item.get("Portal", "Income Tax"), item.get("name", ""),
            item.get("pan", ""), item.get("gstin", ""), item.get("form", ""), item.get("period", ""),
            item.get("arn", "N/A"), item.get("status", "Filed & Verified")
        ])
        for c in range(1, len(headers_cat1) + 1):
            cell = ws_cat1.cell(row=r_idx, column=c)
            cell.border = THIN_BORDER
            cell.font = Font(name="Segoe UI", size=9.5)
            if c == 9: # ARN
                cell.font = Font(name="Consolas", size=10, bold=True, color="008000")
                cell.alignment = Alignment(horizontal="center")

    # --- TAB 3: SUBMITTED PENDING (CAT 2) ---
    ws_cat2 = wb.create_sheet(title="2. Submitted (Pending)")
    ws_cat2.views.sheetView[0].showGridLines = True
    headers_cat2 = ["Entry #", "Timestamp", "Portal", "Assessee Name", "PAN", "Form", "Period", "Stage / Action", "URL Route"]
    ws_cat2.append(headers_cat2)
    for col in range(1, len(headers_cat2) + 1):
        cell = ws_cat2.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, item in enumerate(classified["cat2"], start=2):
        ws_cat2.append([
            item.get("entry_num", ""), item.get("timestamp", ""), item.get("Portal", "Income Tax"), item.get("name", ""),
            item.get("pan", ""), item.get("form", ""), item.get("period", ""), item.get("confirmation_text") or "ITR-V Download / e-Verify Later",
            item.get("url", "")
        ])
        for c in range(1, len(headers_cat2) + 1):
            cell = ws_cat2.cell(row=r_idx, column=c)
            cell.border = THIN_BORDER
            cell.font = Font(name="Segoe UI", size=9.5)

    # --- TAB 4: SDC SESSION TIMELINES (CAT 3) ---
    ws_cat3 = wb.create_sheet(title="3. SDC Session Timelines")
    ws_cat3.views.sheetView[0].showGridLines = True
    headers_cat3 = ["Session ID", "Step #", "Timestamp", "Client Name", "PAN", "Portal", "Session Status", "Action / Route Title", "Portal URL", "Crosshair Trigger", "Captured Parameters"]
    ws_cat3.append(headers_cat3)
    for col in range(1, len(headers_cat3) + 1):
        cell = ws_cat3.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, item in enumerate(classified["cat3"], start=2):
        ws_cat3.append([
            item.get("session_id", ""), item.get("step", 1), item.get("timestamp", ""),
            item.get("client_name", ""), item.get("pan", ""), item.get("portal", ""),
            item.get("session_status", "completed").upper(), item.get("action_title", ""),
            item.get("url", ""), item.get("crosshair_id", ""), item.get("parameters", "-")
        ])
        for c in range(1, len(headers_cat3) + 1):
            cell = ws_cat3.cell(row=r_idx, column=c)
            cell.border = THIN_BORDER
            cell.font = Font(name="Segoe UI", size=9.5)
            if c == 1:
                cell.font = Font(name="Consolas", size=9, bold=True, color="1A365D")
            elif c == 7: # Session Status
                st = item.get("session_status", "")
                if st == "completed":
                    cell.fill = COLORS["status_completed"]
                    cell.font = Font(name="Segoe UI", size=9, bold=True, color="008000")
                elif st == "terminated_abruptly":
                    cell.fill = COLORS["status_abrupt"]
                    cell.font = Font(name="Segoe UI", size=9, bold=True, color="C53030")
                else:
                    cell.fill = COLORS["status_active"]
                    cell.font = Font(name="Segoe UI", size=9, bold=True, color="D97706")
                cell.alignment = Alignment(horizontal="center")

    # --- TAB 5: ACTIVE DRAFTS & SCHEDULES (CAT 4) ---
    ws_cat4 = wb.create_sheet(title="4. Return Drafts & Schedules")
    ws_cat4.views.sheetView[0].showGridLines = True
    headers_cat4 = ["Entry #", "Timestamp", "Portal", "Assessee Name", "PAN", "Form", "Period", "Active Section / Breadcrumbs", "URL Route"]
    ws_cat4.append(headers_cat4)
    for col in range(1, len(headers_cat4) + 1):
        cell = ws_cat4.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, item in enumerate(classified["cat4"], start=2):
        ws_cat4.append([
            item.get("entry_num", ""), item.get("timestamp", ""), item.get("Portal", "Income Tax"), item.get("name", ""),
            item.get("pan", ""), item.get("form", ""), item.get("period", ""), item.get("breadcrumbs") or "Schedule Editing",
            item.get("url", "")
        ])
        for c in range(1, len(headers_cat4) + 1):
            cell = ws_cat4.cell(row=r_idx, column=c)
            cell.border = THIN_BORDER
            cell.font = Font(name="Segoe UI", size=9.5)

    # --- TAB 6: TAXPAYER IDENTITY LEDGER (CAT 5) ---
    ws_cat5 = wb.create_sheet(title="5. Taxpayer Identity Ledger")
    ws_cat5.views.sheetView[0].showGridLines = True
    headers_cat5 = ["PAN", "GSTIN", "Reconciled Taxpayer Name", "Forms Observed", "Latest Period", "Total DOM Captures"]
    ws_cat5.append(headers_cat5)
    for col in range(1, len(headers_cat5) + 1):
        cell = ws_cat5.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, (k, p_obj) in enumerate(classified["cat5"].items(), start=2):
        ws_cat5.append([
            p_obj["pan"], p_obj["gstin"], p_obj["name"],
            ", ".join(sorted(p_obj["forms_seen"])), p_obj["latest_period"],
            p_obj["total_captures"]
        ])
        for c in range(1, len(headers_cat5) + 1):
            cell = ws_cat5.cell(row=r_idx, column=c)
            cell.border = THIN_BORDER
            cell.font = Font(name="Segoe UI", size=9.5)
            if c == 1:
                cell.font = Font(name="Consolas", size=10, bold=True, color="1A365D")
            elif c == 6:
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(name="Segoe UI", size=10, bold=True)

    # --- TAB 7: NAVIGATION JOURNEYS (CAT 6) ---
    ws_cat6 = wb.create_sheet(title="6. Navigation Journey Trace")
    ws_cat6.views.sheetView[0].showGridLines = True
    headers_cat6 = ["Entry #", "Timestamp", "Assessee Name", "PAN", "Portal Route / Breadcrumbs", "Step-by-Step Link History"]
    ws_cat6.append(headers_cat6)
    for col in range(1, len(headers_cat6) + 1):
        cell = ws_cat6.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, item in enumerate(classified["cat6"], start=2):
        ws_cat6.append([
            item.get("entry_num", ""), item.get("timestamp", ""), item.get("name", ""), item.get("pan", ""),
            item.get("breadcrumbs", ""), item.get("link_history", "")
        ])
        for c in range(1, len(headers_cat6) + 1):
            cell = ws_cat6.cell(row=r_idx, column=c)
            cell.border = THIN_BORDER
            cell.font = Font(name="Segoe UI", size=9.5)
            if c == 6:
                cell.font = Font(name="Consolas", size=8.5)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-adjust column widths across all sheets
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 65)

    # Ensure target directory exists
    out_dir = os.path.dirname(output_path)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    try:
        wb.save(output_path)
        print(f"[DOM_Parser_1] Successfully generated audit report: {output_path}")
    except PermissionError:
        fallback_path = output_path.replace(".xlsx", "_latest.xlsx")
        try:
            wb.save(fallback_path)
            print(f"[DOM_Parser_1] Notice: {os.path.basename(output_path)} is open in Excel. Saved updated report to: {fallback_path}")
        except Exception as e:
            print(f"[DOM_Parser_1] Could not save Excel report (file locked in Excel): {e}")
    return True

def process_data(dump_path=None, report_path=None):
    """Main programmatic interface for Project Sera."""
    if not report_path:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        report_path = os.path.join(base_dir, "dom_audit_report.xlsx")
    target_dump = resolve_target_dump(dump_path)
    entries, timelines = load_data(target_dump)
    classified = classify_entries(entries, timelines)
    return generate_excel_report(classified, report_path)

def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    default_report = os.path.join(base_dir, "dom_audit_report.xlsx")
    parser = argparse.ArgumentParser(description="DOM_Parser_1: Analytical Classifier & Audit Report Generator for Sera DOM")
    parser.add_argument("input", nargs="?", default="rawPayload.db", help="Path to rawPayload.db or seraRawPayloadDump.txt")
    parser.add_argument("output", nargs="?", default=default_report, help="Output Excel report path")
    parser.add_argument("--watch", action="store_true", help="Continuously watch for dump updates and re-generate report")
    args = parser.parse_args()

    input_file = resolve_target_dump(args.input)
    print(f"=== DOM_Parser_1 (Sera Visual Layer Classifier) ===")
    print(f"Target Input:  {input_file}")
    print(f"Output Report: {args.output}")

    if args.watch:
        print(f"[DOM_Parser_1] Watching for updates on {input_file} (Polling every 3s)...")
        last_mtime = 0
        while True:
            try:
                if os.path.exists(input_file):
                    mtime = os.path.getmtime(input_file)
                    if mtime != last_mtime:
                        last_mtime = mtime
                        process_data(input_file, args.output)
            except Exception as e:
                print(f"[DOM_Parser_1] Watcher error: {e}")
            time.sleep(3)
    else:
        process_data(input_file, args.output)

if __name__ == "__main__":
    main()
