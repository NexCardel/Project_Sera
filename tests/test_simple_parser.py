import os
import tempfile
import unittest

from simpleParser.simple_parser import _event, process_dump


ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
MOCK = os.path.join(ROOT, "seraRawPayloadDump_mock.txt")


class TestSimpleParser(unittest.TestCase):
    def test_gst_nested_business_and_legal_names(self):
        event = _event({
            "entry_number": 1,
            "timestamp": "2026-08-26T10:00:00+00:00",
            "portal": "GST Portal",
            "payload": {
                "pan": "AIUPA2571J",
                "raw_payload": {
                    "data": {
                        "bn": "MOHAMMAD NOOR AZAM MOLLA",
                        "ln": "MOHAMMAD NOOR AZAM MOLLA",
                        "tn": "A NOOR AZAM DRESSES",
                        "gstin": "19AIUPA2571J1ZW",
                    }
                },
                "url": "auth/api/formdetails?rtn_prd=062026&rtn_typ=GSTR3B",
            },
        })
        self.assertIn({"value": "MOHAMMAD NOOR AZAM MOLLA", "source": "raw_payload.data.bn", "role": "business_name"}, event["names"])
        self.assertIn({"value": "MOHAMMAD NOOR AZAM MOLLA", "source": "raw_payload.data.ln", "role": "legal_name"}, event["names"])
        self.assertIn({"value": "A NOOR AZAM DRESSES", "source": "raw_payload.data.tn", "role": "business_name"}, event["names"])

    def test_tracker_grouped_views_stack_in_opposite_vertical_directions(self):
        from simpleParser.simple_parser import _tracker_grouped_rows
        events = [
            {"entry": 1, "timestamp": "2026-01-01T10:00:00", "pans": ["ABCDE1234F"], "names": [], "transactions": []},
            {"entry": 2, "timestamp": "2026-01-01T11:00:00", "pans": ["ABCDE1234F"], "names": [], "transactions": []},
        ]
        _, up_rows, _, down_rows = _tracker_grouped_rows(events)
        self.assertEqual([row[3] for row in up_rows], [2, 1])
        self.assertEqual([row[3] for row in down_rows], [1, 2])

    def test_transaction_linked_submission_and_everification(self):
        result = process_dump(MOCK)
        row = next(row for row in result["lifecycle_rows"] if row[0] == "MOKPA1000A")
        self.assertEqual(row[1], "Return submitted and e-verified")
        self.assertIn("ITR000883649705", row[7])
        self.assertIn("EVERIFY000920870466", row[8])

    def test_report_contains_expected_sheets(self):
        with tempfile.TemporaryDirectory() as temp:
            report = os.path.join(temp, "simple_parser_report.xlsx")
            result = process_dump(MOCK, output_excel=report, master_pans={"MOKPA1000A"})
            self.assertTrue(os.path.exists(result["outputs"]["excel_path"]))
            from openpyxl import load_workbook
            workbook = load_workbook(result["outputs"]["excel_path"], read_only=True)
            self.assertEqual(
                set(workbook.sheetnames),
                {"Lifecycle Summary", "trackerDumpGroupedUpStacked", "trackerDumpGroupedDownStacked", "Parser Events", "Identity Evidence", "Quarantine"},
            )
            summary = workbook["Lifecycle Summary"]
            self.assertEqual(summary[1][0].value, "PAN")
            self.assertEqual(summary[2][0].value, "MOKPA1000A")
            self.assertEqual(summary[2][9].value, "Yes")
            workbook.close()


if __name__ == "__main__":
    unittest.main()
