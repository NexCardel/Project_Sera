"""FST Tracer Alpha: conservative PAN/GSTIN-first payload analysis.

This is intentionally independent of ``tracker_dump_parser`` and
``FST_Classifier_1``.  It treats the raw dump as an evidence ledger, builds
client containers keyed by PAN (or GSTIN when PAN cannot be derived), groups
events into bounded sessions, and emits a human-readable Excel workbook.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import time
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

PAN_RE = re.compile(r"\b[A-Z]{5}[0-9]{4}[A-Z]\b", re.I)
GSTIN_RE = re.compile(r"\b[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][1-9A-Z]Z[0-9A-Z]\b", re.I)
ENTRY_RE = re.compile(r"CAPTURE DUMP ENTRY #\s*(\d+)")
SEPARATOR = "========================================================================================"
DEFAULT_CONTEXT_SECONDS = 15 * 60
DEFAULT_SESSION_GAP_SECONDS = 20 * 60
DEFAULT_OBSIDIAN_VAULT = Path(__file__).resolve().parent.parent / "docs" / "APP" / "Sera FST Tracer Alpha"


def _dt(value: Any) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    except (TypeError, ValueError):
        return None


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    value = re.sub(r"\s+", " ", str(value)).strip()
    return value or None


def _walk(value: Any) -> Iterable[tuple[str, Any]]:
    if isinstance(value, dict):
        for key, child in value.items():
            yield str(key), child
            yield from _walk(child)
    elif isinstance(value, list):
        for child in value:
            yield from _walk(child)


def _all_strings(value: Any) -> Iterable[str]:
    if isinstance(value, str):
        yield value
    elif isinstance(value, dict):
        for child in value.values():
            yield from _all_strings(child)
    elif isinstance(value, list):
        for child in value:
            yield from _all_strings(child)


def _valid_pan(value: Any) -> str | None:
    match = PAN_RE.fullmatch(str(value).strip().upper()) if value is not None else None
    return match.group(0) if match else None


def _valid_gstin(value: Any) -> str | None:
    match = GSTIN_RE.fullmatch(str(value).strip().upper()) if value is not None else None
    return match.group(0) if match else None


def _identity_candidates(header: dict[str, Any], envelope: dict[str, Any], raw: dict[str, Any]) -> list[dict[str, str]]:
    candidates: list[dict[str, str]] = []

    def add(value: Any, source: str, kind: str = "PAN") -> None:
        if kind == "PAN":
            value = _valid_pan(value)
        else:
            value = _valid_gstin(value)
        if value and not any(c["value"] == value for c in candidates):
            candidates.append({"value": value, "kind": kind, "source": source})

    add(header.get("header_pan"), "header")
    for key in ("pan", "panNumber", "pan_no", "entityNum", "submitUserId", "loggedInUserId", "userId"):
        add(envelope.get(key), f"envelope.{key}")
        add(raw.get(key), f"raw_payload.{key}")
    for key in ("gstin", "gstinNo", "gstin_num"):
        add(envelope.get(key), f"envelope.{key}", "GSTIN")
        add(raw.get(key), f"raw_payload.{key}", "GSTIN")

    # GST profile tokens are commonly carried in synthetic ACK/profile labels.
    for text in list(_all_strings(envelope)) + list(_all_strings(raw)):
        for gstin in GSTIN_RE.findall(text.upper()):
            add(gstin, "deep.GSTIN", "GSTIN")
        for pan in PAN_RE.findall(text.upper()):
            add(pan, "deep.PAN")

    # A GSTIN contains the PAN at positions 3..12 (zero-based 2:12).
    for candidate in list(candidates):
        if candidate["kind"] == "GSTIN":
            pan = _valid_pan(candidate["value"][2:12])
            if pan and not any(c["value"] == pan for c in candidates):
                candidates.append({"value": pan, "kind": "PAN", "source": candidate["source"] + ".derived_pan"})
    return candidates


def _name(envelope: dict[str, Any], raw: dict[str, Any]) -> str | None:
    for obj in (raw, envelope):
        for key in ("nameAsPerBank", "bn", "auth_name", "assesseName", "assesseeName"):
            value = _clean(obj.get(key))
            if value:
                return value
        parts = [_clean(obj.get(k)) for k in ("firstName", "midName", "middleName", "lastName")]
        parts = [p for p in parts if p and p.lower() != "none"]
        if parts:
            return " ".join(parts)
        for _, child in _walk(obj):
            if isinstance(child, dict):
                an = child.get("AssesseeName")
                if isinstance(an, dict):
                    value = " ".join(filter(None, (_clean(an.get(k)) for k in ("FirstName", "MiddleName", "SurNameOrOrgName"))))
                    if value:
                        return value
    return None


def _header(chunk: str) -> dict[str, Any]:
    result: dict[str, Any] = {}
    match = ENTRY_RE.search(chunk)
    result["entry_number"] = int(match.group(1)) if match else None
    for line in chunk.splitlines():
        if ":" not in line or line.strip().startswith("RAW JSON"):
            continue
        key, value = line.split(":", 1)
        key = key.strip().lower().replace(" ", "_").replace("/", "_")
        result[key] = value.strip()
    client_text = result.get("client_id", "")
    pan = PAN_RE.search(client_text.upper())
    result["header_pan"] = pan.group(0) if pan else None
    return result


def parse_dump(source: str | Path) -> list[dict[str, Any]]:
    """Parse dump blocks without dropping malformed blocks."""
    path = Path(source)
    text = path.read_text(encoding="utf-8", errors="replace") if path.is_file() else str(source)
    chunks = re.split(r"(?=CAPTURE DUMP ENTRY #)", text)
    entries: list[dict[str, Any]] = []
    for chunk in chunks:
        if not ENTRY_RE.search(chunk):
            continue
        item = _header(chunk)
        marker = "RAW JSON PAYLOAD:"
        body = chunk.split(marker, 1)[1] if marker in chunk else ""
        body = body.split(SEPARATOR, 1)[0].strip()
        try:
            item["payload"] = json.loads(body)
            item["parse_error"] = None
        except json.JSONDecodeError as exc:
            item["payload"] = None
            item["parse_error"] = f"JSON decode error: {exc.msg} at character {exc.pos}"
        entries.append(item)
    return entries


def _ack(item: dict[str, Any], envelope: dict[str, Any], raw: dict[str, Any]) -> str | None:
    for value in (item.get("arn___ack_no"), envelope.get("arn"), raw.get("arnNumber"), raw.get("ackNum"), raw.get("transactionNo")):
        value = _clean(value)
        if value and value not in {"N/A", "null", "None"}:
            return value
    return None


def _form_type(envelope: dict[str, Any], raw: dict[str, Any]) -> str | None:
    for key in ("filing_type", "filingTypeCd", "formTypeCd", "formCd", "rtn_type", "formName"):
        value = _clean(envelope.get(key) or raw.get(key))
        if value:
            value = value.upper()
            if value.isdigit():
                return "ITR-" + value
            return value
    url = str(envelope.get("url", "")).lower()
    match = re.search(r"rtn_typ=([A-Z0-9-]+)", url, re.I)
    return match.group(1).upper() if match else None


def _period_label(item: dict[str, Any], envelope: dict[str, Any], raw: dict[str, Any]) -> str | None:
    """Extract a filing period from headers and common portal payload keys."""
    for obj in (item, envelope, raw):
        for key in ("filing_period", "period_label", "periodLabel", "ret_period", "ret_prd", "period"):
            value = _clean(obj.get(key)) if isinstance(obj, dict) else None
            if value and value.lower() not in {"n/a", "na", "null", "none"}:
                return value
    return None


def _decode_action(envelope: dict[str, Any], raw: dict[str, Any]) -> tuple[str, str, str]:
    url = str(envelope.get("url", "")).lower()
    status = str(raw.get("status") or raw.get("status_cd") or envelope.get("status") or "").upper()
    if "validateotp" in url:
        module = str(raw.get("moduleCode", "")).upper()
        return ("ITR e-verification" if module == "ITR" else "Bank e-verification", "e-verified" if status in {"SUCCESS", "1"} else "verification-failed", "OTP validation response")
    if "submit/wzrd" in url:
        accepted = str(raw.get("httpStatus", "")).upper() == "ACCEPTED" or raw.get("successFlag") is True
        return ("Return submission", "submitted" if accepted else "submission-attempt", "wizard submission response")
    if "formdetails" in url:
        return ("GST return filing", "filed" if status in {"FIL", "1", "SUCCESS"} else "filing-review", "GST form response")
    if "getentity" in url and any(k in raw for k in ("bankName", "accValidity", "accountStatus")):
        return ("Bank account validation/status", "validated" if raw.get("accValidity") == "V" else "review", "bank response")
    if "saveentity" in url:
        return ("Profile/entity update", "saved", "profile response")
    if "login" in url:
        return ("Portal login", "success" if status in {"SUCCESS", "1"} else "login-response", "authentication response")
    if "downloadfile" in url:
        return ("Filed return download", "downloaded", "return document response")
    if "details" in url:
        return ("Return details", "retrieved", "return detail response")
    if "summary" in url or "filingsnapshot" in url:
        return ("GST return history/summary", "retrieved", "GST history response")
    return ("Portal interaction", "observed", "unclassified endpoint")


def _event(item: dict[str, Any]) -> dict[str, Any]:
    envelope = item.get("payload") or {}
    raw = envelope.get("raw_payload") if isinstance(envelope, dict) else {}
    if isinstance(raw, str):
        try:
            raw = json.loads(raw)
        except json.JSONDecodeError:
            raw = {}
    raw = raw if isinstance(raw, dict) else {}
    candidates = _identity_candidates(item, envelope, raw)
    pan_values = sorted({c["value"] for c in candidates if c["kind"] == "PAN"})
    gstin_values = sorted({c["value"] for c in candidates if c["kind"] == "GSTIN"})
    action, outcome, reason = _decode_action(envelope, raw)
    return {
        "entry_number": item.get("entry_number"),
        "timestamp": item.get("timestamp"),
        "timestamp_dt": _dt(item.get("timestamp")),
        "portal": item.get("portal") or envelope.get("portal"),
        "capture_method": item.get("capture_method"),
        "url": envelope.get("url", ""),
        "action": action,
        "outcome": outcome,
        "reason": reason,
        "filing_type": _form_type(envelope, raw),
        "period_label": _period_label(item, envelope, raw),
        "ack": _ack(item, envelope, raw),
        "pan_candidates": pan_values,
        "gstin_candidates": gstin_values,
        "identity_sources": [c["source"] for c in candidates],
        "name": _name(envelope, raw),
        "envelope": envelope,
        "raw_payload": raw,
        "parse_error": item.get("parse_error"),
    }


def _key_for(event: dict[str, Any]) -> str | None:
    pans = event["pan_candidates"]
    gstins = event["gstin_candidates"]
    if len(pans) == 1:
        return "PAN:" + pans[0]
    if len(pans) > 1:
        return None
    if len(gstins) == 1:
        return "GSTIN:" + gstins[0]
    return None


def trace_payloads(entries: list[dict[str, Any]], context_seconds: int = DEFAULT_CONTEXT_SECONDS, session_gap_seconds: int = DEFAULT_SESSION_GAP_SECONDS) -> dict[str, Any]:
    events = [_event(item) for item in entries]
    ack_to_key: dict[str, str] = {}
    for event in events:
        key = _key_for(event)
        if key and event["ack"]:
            ack_to_key.setdefault(event["ack"], key)

    anchored = [(i, e, _key_for(e)) for i, e in enumerate(events) if _key_for(e)]
    for index, event in enumerate(events):
        key = _key_for(event)
        if key:
            event.update(client_key=key, identity_method="explicit", identity_confidence="high", identity_status="resolved")
            continue
        if event["ack"] in ack_to_key:
            event.update(client_key=ack_to_key[event["ack"]], identity_method="ack-link", identity_confidence="high", identity_status="resolved")
            continue
        current_time = event["timestamp_dt"]
        nearby = []
        if current_time:
            for anchor_index, anchor, anchor_key in anchored:
                if not anchor["timestamp_dt"] or anchor["portal"] != event["portal"]:
                    continue
                distance = abs((current_time - anchor["timestamp_dt"]).total_seconds())
                if distance <= context_seconds:
                    nearby.append((distance, anchor_key))
        keys = {key for _, key in nearby}
        if len(keys) == 1:
            event.update(client_key=next(iter(keys)), identity_method="bounded-context", identity_confidence="medium", identity_status="resolved")
        elif len(keys) > 1:
            event.update(client_key=None, identity_method="conflicting-context", identity_confidence="none", identity_status="ambiguous")
        else:
            event.update(client_key=None, identity_method="no-evidence", identity_confidence="none", identity_status="unresolved")

    clients: dict[str, dict[str, Any]] = {}
    quarantine: list[dict[str, Any]] = []
    for event in events:
        if event.get("client_key"):
            client = clients.setdefault(event["client_key"], {"client_key": event["client_key"], "pans": set(), "gstins": set(), "names": set(), "events": []})
            client["events"].append(event)
            client["pans"].update(event["pan_candidates"])
            client["gstins"].update(event["gstin_candidates"])
            if event["name"]:
                client["names"].add(event["name"])
        else:
            quarantine.append({"entry_number": event["entry_number"], "status": event["identity_status"], "method": event["identity_method"], "ack": event["ack"], "url": event["url"], "reason": "No single defensible PAN/GSTIN owner"})

    sessions: list[dict[str, Any]] = []
    for client in clients.values():
        client["events"].sort(key=lambda e: e["timestamp_dt"] or datetime.min.replace(tzinfo=timezone.utc))
        current: list[dict[str, Any]] = []
        last_dt = None
        for event in client["events"]:
            dt = event["timestamp_dt"]
            gap = (dt - last_dt).total_seconds() if dt and last_dt else 0
            if current and (gap > session_gap_seconds or event["portal"] != current[-1]["portal"]):
                sessions.append(_make_session(client["client_key"], len([s for s in sessions if s["client_key"] == client["client_key"]]) + 1, current))
                current = []
            current.append(event)
            if dt:
                last_dt = dt
        if current:
            sessions.append(_make_session(client["client_key"], len([s for s in sessions if s["client_key"] == client["client_key"]]) + 1, current))

    for client in clients.values():
        client["pans"] = sorted(client["pans"])
        client["gstins"] = sorted(client["gstins"])
        client["names"] = sorted(client["names"])
        client["event_count"] = len(client["events"])
        client["session_count"] = sum(1 for s in sessions if s["client_key"] == client["client_key"])
        client["filing_events"] = [e for e in client["events"] if e["action"] in {"Return submission", "ITR e-verification", "Bank e-verification", "GST return filing"}]
        del client["events"]

    return {"entries": events, "clients": clients, "sessions": sessions, "quarantine": quarantine, "stats": {"input_entries": len(entries), "resolved_entries": len(events) - len(quarantine), "quarantine_entries": len(quarantine), "client_containers": len(clients), "sessions": len(sessions)}}


def _make_session(client_key: str, number: int, events: list[dict[str, Any]]) -> dict[str, Any]:
    start, end = events[0]["timestamp_dt"], events[-1]["timestamp_dt"]
    return {"client_key": client_key, "session_number": number, "start": events[0]["timestamp"], "end": events[-1]["timestamp"], "duration_seconds": int((end - start).total_seconds()) if start and end else 0, "event_count": len(events), "events": events}


def _parser_data_rows(result: dict[str, Any], master_pans: Iterable[str] | None = None) -> list[list[Any]]:
    """Create the human-facing extraction table requested for the live report."""
    master = {str(p).strip().upper() for p in (master_pans or []) if _valid_pan(p)}
    rows: list[list[Any]] = []
    filing_actions = {"Return submission", "ITR e-verification", "Bank e-verification", "GST return filing"}

    for key, client in sorted(result["clients"].items()):
        events = [e for e in result["entries"] if e.get("client_key") == key]
        events.sort(key=lambda e: e.get("timestamp_dt") or datetime.min.replace(tzinfo=timezone.utc))
        filing_events = [e for e in events if e.get("action") in filing_actions]
        latest = events[-1] if events else {}
        latest_filing = filing_events[-1] if filing_events else latest
        pans = list(client.get("pans", []))
        pan = ", ".join(pans) or ""

        has_submission = any(e.get("action") in {"Return submission", "GST return filing"} and e.get("outcome") in {"submitted", "filed"} for e in events)
        has_itr_everification = any(e.get("action") == "ITR e-verification" and e.get("outcome") == "e-verified" for e in events)
        has_bank_everification = any(e.get("action") == "Bank e-verification" and e.get("outcome") == "e-verified" for e in events)
        if has_submission and (has_itr_everification or has_bank_everification):
            conclusion = "Return submitted and e-verified"
        elif has_submission:
            conclusion = "Return submitted; not e-verified"
        elif has_itr_everification or has_bank_everification:
            conclusion = "EVC/e-verification event detected; no return submission"
        else:
            conclusion = f"{latest.get('action') or 'No action'} — {latest.get('outcome') or 'unclassified'}"

        rows.append([
            pan,
            "; ".join(client.get("names", [])) or key,
            latest_filing.get("filing_type") or "",
            latest_filing.get("period_label") or "",
            latest.get("action") or "",
            conclusion,
            "Yes" if any(p in master for p in pans) else ("Not parsed" if not pans else "No"),
        ])
    return rows


def _write_excel(result: dict[str, Any], output: str | Path, master_pans: Iterable[str] | None = None) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="203864")
    header_font = Font(color="FFFFFF", bold=True)

    def sheet(name: str, headers: list[str]):
        ws = wb.create_sheet(name) if len(wb.worksheets) > 1 or wb.active.max_row > 1 else wb.active
        ws.title = name
        ws.append(headers)
        for cell in ws[1]:
            cell.fill, cell.font, cell.alignment = header_fill, header_font, Alignment(wrap_text=True)
        ws.freeze_panes = "A2"
        return ws

    ws = sheet("Client Containers", ["Client Key", "PAN(s)", "GSTIN(s)", "Client Name(s)", "Events", "Sessions", "Filing Events"])
    for key, c in sorted(result["clients"].items()):
        ws.append([key, ", ".join(c["pans"]), ", ".join(c["gstins"]), "; ".join(c["names"]), c["event_count"], c["session_count"], len(c["filing_events"])])

    ws = wb.create_sheet("Parser Data Extraction")
    ws.append(["PAN", "Client Name / Parsed Identity", "Filing Type / Form Type", "Filing Period", "Last Action", "Submission Conclusion", "PAN Present in Master DB"])
    for cell in ws[1]: cell.fill, cell.font = header_fill, header_font
    for row in _parser_data_rows(result, master_pans):
        ws.append(row)

    ws = wb.create_sheet("Session Timelines")
    ws.append(["Client Key", "Session", "Start", "End", "Duration (sec)", "Events", "Human Timeline"])
    for cell in ws[1]: cell.fill, cell.font = header_fill, header_font
    for s in result["sessions"]:
        lines = [f"{e['timestamp']} — {e['action']} — {e['outcome']}" + (f" — {e['filing_type']}" if e["filing_type"] else "") + (f" — ACK {e['ack']}" if e["ack"] else "") for e in s["events"]]
        ws.append([s["client_key"], s["session_number"], s["start"], s["end"], s["duration_seconds"], s["event_count"], "\n".join(lines)])

    ws = wb.create_sheet("Filing Events")
    ws.append(["Client Key", "Entry", "Timestamp", "Action", "Filing Type", "Outcome", "ACK/ARN", "Identity", "Endpoint"])
    for cell in ws[1]: cell.fill, cell.font = header_fill, header_font
    for e in result["entries"]:
        if e["action"] in {"Return submission", "ITR e-verification", "Bank e-verification", "GST return filing"}:
            ws.append([e.get("client_key"), e["entry_number"], e["timestamp"], e["action"], e["filing_type"], e["outcome"], e["ack"], e["identity_method"], e["url"]])

    ws = wb.create_sheet("Event Ledger")
    ws.append(["Entry", "Timestamp", "Client Key", "Identity Status", "Identity Method", "Portal", "Action", "Outcome", "ACK/ARN", "Endpoint"])
    for cell in ws[1]: cell.fill, cell.font = header_fill, header_font
    for e in result["entries"]:
        ws.append([e["entry_number"], e["timestamp"], e.get("client_key"), e["identity_status"], e["identity_method"], e["portal"], e["action"], e["outcome"], e["ack"], e["url"]])

    ws = wb.create_sheet("Quarantine")
    ws.append(["Entry", "Status", "Method", "ACK/ARN", "Endpoint", "Reason"])
    for cell in ws[1]: cell.fill, cell.font = header_fill, header_font
    for q in result["quarantine"]:
        ws.append([q["entry_number"], q["status"], q["method"], q["ack"], q["url"], q["reason"]])

    for ws in wb.worksheets:
        ws.auto_filter.ref = ws.dimensions
        for column in ws.columns:
            letter = column[0].column_letter
            ws.column_dimensions[letter].width = min(max(max(len(str(c.value or "")) for c in column) + 2, 12), 55)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    output = Path(output)
    output.parent.mkdir(parents=True, exist_ok=True)
    temp_output = output.with_name(f".{output.stem}.{os.getpid()}.tmp.xlsx")
    try:
        # Save a complete workbook first, then replace the published report.
        # This prevents readers from seeing a partially written workbook.
        wb.save(temp_output)
        try:
            os.replace(temp_output, output)
            return output
        except PermissionError:
            # Excel holds open files with an exclusive sharing lock on Windows.
            # Publish a fresh sibling instead of dropping the live refresh.
            latest = output.with_name(f"{output.stem}_latest{output.suffix}")
            try:
                os.replace(temp_output, latest)
                return latest
            except PermissionError:
                timestamped = output.with_name(f"{output.stem}_{time.strftime('%Y%m%d_%H%M%S')}_{time.time_ns() % 1000000:06d}{output.suffix}")
                os.replace(temp_output, timestamped)
                return timestamped
    finally:
        try:
            wb.close()
        except Exception:
            pass
        if temp_output.exists():
            try:
                temp_output.unlink()
            except OSError:
                pass


def _md_safe(value: Any) -> str:
    value = re.sub(r"[^A-Za-z0-9._-]+", "_", str(value or "UNKNOWN"))
    return value.strip("._") or "UNKNOWN"


def _md_text(value: Any) -> str:
    return str(value or "").replace("|", "\\|").replace("\r", "").strip()


def _client_label(client_key: str, client: dict[str, Any]) -> str:
    identity = ", ".join(client.get("pans") or client.get("gstins") or [client_key])
    names = "; ".join(client.get("names", []))
    return f"{identity} — {names}" if names else identity


def _write_obsidian_vault(result: dict[str, Any], vault_root: str | Path) -> None:
    """Write the generated FST timeline notes into an Obsidian-compatible vault.

    Only the application-owned ``Sera FST Tracer Alpha`` folder is replaced on
    refresh, and only when its marker file is present. This prevents an
    accidentally pointed-at personal vault from being deleted.
    """
    root = Path(vault_root)
    marker = root / ".sera-fst-tracer-alpha-generated"
    if root.exists() and marker.exists():
        shutil.rmtree(root)
    root.mkdir(parents=True, exist_ok=True)
    marker.write_text("Project Sera generated folder. Do not edit manually.\n", encoding="utf-8")
    (root / "Clients").mkdir(exist_ok=True)
    (root / "Sessions").mkdir(exist_ok=True)
    (root / "Filing Events").mkdir(exist_ok=True)
    (root / "Quarantine").mkdir(exist_ok=True)

    stats = result["stats"]
    client_links = []
    for key, client in sorted(result["clients"].items()):
        filename = _md_safe(key) + ".md"
        client_links.append(f"- [[Clients/{filename[:-3]}|{_client_label(key, client)}]]")

    dashboard = [
        "# Sera FST Tracer Alpha",
        "",
        "> Auto-generated Obsidian vault for evidence-first raw API timelines.",
        "> Refresh this folder from the Sera Tracker Dump pipeline; do not edit generated notes.",
        "",
        f"- Generated: `{datetime.now(timezone.utc).isoformat()}`",
        f"- Input entries: **{stats['input_entries']}**",
        f"- Resolved entries: **{stats['resolved_entries']}**",
        f"- Client containers: **{stats['client_containers']}**",
        f"- Sessions: **{stats['sessions']}**",
        f"- Quarantined entries: **{stats['quarantine_entries']}**",
        "",
        "## Client containers",
        "",
        *client_links,
        "",
        "## Other views",
        "",
        "- [[Quarantine]]",
        "- [[Filing Events Index]]",
        "- [[Session Index]]",
    ]
    (root / "Sera FST Tracer Alpha.md").write_text("\n".join(dashboard) + "\n", encoding="utf-8")

    session_links = []
    for session in result["sessions"]:
        session_file = f"{_md_safe(session['client_key'])}__S{session['session_number']:02d}.md"
        session_links.append(f"- [[Sessions/{session_file[:-3]}|{session['client_key']} — Session {session['session_number']:02d}]]")
        lines = [
            f"# {session['client_key']} — Session {session['session_number']:02d}",
            "",
            f"- Client: [[Clients/{_md_safe(session['client_key'])}|{session['client_key']}]]",
            f"- Start: `{session['start']}`",
            f"- End: `{session['end']}`",
            f"- Duration: `{session['duration_seconds']} seconds`",
            f"- Events: `{session['event_count']}`",
            "",
            "## Human-readable timeline",
            "",
        ]
        for event in session["events"]:
            filing = f" — {event['filing_type']}" if event["filing_type"] else ""
            ack = f" — ACK/ARN `{event['ack']}`" if event["ack"] else ""
            lines.append(f"- `{event['timestamp']}` — **{_md_text(event['action'])}** — {_md_text(event['outcome'])}{filing}{ack} (entry #{event['entry_number']})")
        (root / "Sessions" / session_file).write_text("\n".join(lines) + "\n", encoding="utf-8")
    (root / "Session Index.md").write_text("# Session Index\n\n" + "\n".join(session_links) + "\n", encoding="utf-8")

    filing_links = []
    filing_number = 0
    for event in result["entries"]:
        if event["action"] not in {"Return submission", "ITR e-verification", "Bank e-verification", "GST return filing"}:
            continue
        filing_number += 1
        filename = f"{filing_number:04d}__Entry_{event['entry_number']}__{_md_safe(event['action'])}.md"
        filing_links.append(f"- [[Filing Events/{filename[:-3]}|Entry #{event['entry_number']} — {event['action']}]]")
        lines = [
            f"# Filing Event — Entry #{event['entry_number']}", "",
            f"- Client: [[Clients/{_md_safe(event.get('client_key'))}|{event.get('client_key') or 'Unresolved'}]]",
            f"- Timestamp: `{event['timestamp']}`",
            f"- Action: **{_md_text(event['action'])}**",
            f"- Filing type: `{event['filing_type'] or 'N/A'}`",
            f"- Outcome: **{_md_text(event['outcome'])}**",
            f"- ACK/ARN: `{event['ack'] or 'N/A'}`",
            f"- Identity evidence: `{event['identity_method']}` ({event['identity_confidence']})",
            f"- Endpoint: `{event['url']}`", "",
            "## Traceability", "",
            f"Raw dump entry: `CAPTURE DUMP ENTRY #{event['entry_number']}`",
        ]
        (root / "Filing Events" / filename).write_text("\n".join(lines) + "\n", encoding="utf-8")
    (root / "Filing Events Index.md").write_text("# Filing Events Index\n\n" + "\n".join(filing_links) + "\n", encoding="utf-8")

    quarantine_lines = ["# Quarantine", "", "Events below were not assigned to a client because the evidence was missing or conflicting.", ""]
    for item in result["quarantine"]:
        quarantine_lines.append(f"- Entry #{item['entry_number']} — **{item['status']}** — {item['method']} — ACK/ARN `{item['ack'] or 'N/A'}` — `{item['url']}`")
    (root / "Quarantine.md").write_text("\n".join(quarantine_lines) + "\n", encoding="utf-8")

    for key, client in result["clients"].items():
        session_refs = [s for s in result["sessions"] if s["client_key"] == key]
        lines = [
            f"# {_client_label(key, client)}", "",
            f"- Client key: `{key}`",
            f"- PAN(s): `{', '.join(client['pans']) or 'N/A'}`",
            f"- GSTIN(s): `{', '.join(client['gstins']) or 'N/A'}`",
            f"- Event count: `{client['event_count']}`",
            f"- Session count: `{client['session_count']}`", "",
            "## Sessions", "",
        ]
        for session in session_refs:
            sf = f"{_md_safe(key)}__S{session['session_number']:02d}"
            lines.append(f"- [[Sessions/{sf}|Session {session['session_number']:02d}]] — {session['start']} → {session['end']}")
        lines.extend(["", "## Filing events", ""])
        for event in client["filing_events"]:
            lines.append(f"- Entry #{event['entry_number']}: **{event['action']}** — {event['outcome']} — `{event['ack'] or 'N/A'}`")
        (root / "Clients" / f"{_md_safe(key)}.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def process_dump(input_dump: str | Path, output_excel: str | Path | None = None, output_vault: str | Path | None = None, master_pans: Iterable[str] | None = None) -> dict[str, Any]:
    entries = parse_dump(input_dump)
    result = trace_payloads(entries)
    result["outputs"] = {}
    if output_excel:
        result["outputs"]["excel_path"] = str(_write_excel(result, output_excel, master_pans=master_pans))
    _write_obsidian_vault(result, output_vault or DEFAULT_OBSIDIAN_VAULT)
    result["outputs"]["obsidian_vault"] = str(output_vault or DEFAULT_OBSIDIAN_VAULT)
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Project Sera FST Tracer Alpha")
    parser.add_argument("input_dump", nargs="?", default="..\\seraRawPayloadDump.txt")
    parser.add_argument("output_excel", nargs="?", default="fst_tracer_alpha_report.xlsx")
    parser.add_argument("--vault", default=str(DEFAULT_OBSIDIAN_VAULT), help="Obsidian vault folder for generated timelines")
    parser.add_argument("--watch", action="store_true")
    args = parser.parse_args()
    source = Path(args.input_dump)
    last_mtime = None
    while True:
        if source.exists():
            mtime = source.stat().st_mtime
            if mtime != last_mtime:
                result = process_dump(source, args.output_excel, args.vault)
                print(json.dumps(result["stats"], sort_keys=True))
                last_mtime = mtime
        if not args.watch:
            return
        time.sleep(2)


if __name__ == "__main__":
    main()
