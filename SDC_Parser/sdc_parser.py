import os
import sys
import json
import re
import pandas as pd
from datetime import datetime, date
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DEBUG = True

# Setup paths to import sqlcipher3 and security module independently
APP_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
sys.path.insert(0, APP_DIR)

try:
    import sqlcipher3.dbapi2 as sqlcipher
except ImportError:
    print("Error: sqlcipher3 not found. Make sure to run within the Sera Python environment.")
    sys.exit(1)

def get_db_hex_key():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    app_dir = os.path.abspath(os.path.join(base_dir, ".."))

    salt_candidates = [
        os.path.join(app_dir, "sera.salt"),
        os.path.join(os.path.expanduser("~"), "AmanAssociates_Sera", "sera.salt"),
        os.path.join(app_dir, "salt.bin"),
        os.path.join(os.path.expanduser("~"), "AmanAssociates_Sera", "salt.bin"),
    ]
    key_candidates = [
        os.path.join(app_dir, "sera.key"),
        os.path.join(os.path.expanduser("~"), "AmanAssociates_Sera", "sera.key"),
    ]

    salt = None
    for s_path in salt_candidates:
        if os.path.exists(s_path):
            try:
                with open(s_path, "rb") as f:
                    salt = f.read()
                if salt: break
            except Exception: pass

    pwd = "admin123"
    for k_path in key_candidates:
        if os.path.exists(k_path):
            try:
                with open(k_path, "r", encoding="utf-8") as f:
                    p = f.read().strip()
                if p:
                    pwd = p
                    break
            except Exception: pass
            
    if salt:
        try:
            if app_dir not in sys.path:
                sys.path.insert(0, app_dir)
            import security
            return security.derive_key_hex(pwd, salt)
        except Exception:
            pass
    return None

def get_db_connection():
    # Use the live database in AmanAssociates_Sera
    db_path = os.path.join(os.path.expanduser("~"), "AmanAssociates_Sera", "rawPayload.db")
    if not os.path.exists(db_path):
        db_path = os.path.join(APP_DIR, "rawPayload.db")
        if not os.path.exists(db_path):
            raise FileNotFoundError("rawPayload.db not found.")
            
    hex_key = get_db_hex_key()
    if not hex_key:
        raise ValueError("Could not derive database key")
        
    conn = sqlcipher.connect(db_path)
    conn.execute(f"PRAGMA key = \"x'{hex_key}'\";")
    conn.row_factory = sqlcipher.Row
    return conn

def evaluate_status(raw_status):
    raw = str(raw_status).lower().strip()
    # Strip any legacy bracketed color annotations
    raw = re.sub(r'[\(\[\{]\s*(?:green|yellow|red|blue|gray|grey)\s*[\)\]\}]', '', raw).strip()
    if not raw or raw == "null" or raw == "none":
        return "Not submitted"
    if "not filed" in raw or "unfiled" in raw or "to be filed" in raw:
        return "Not submitted"
    if "pending" in raw:
        return "Submitted (e-verification pending)"
    if "filed" in raw or "portal confirmed" in raw or "verified" in raw:
        return "Submitted & E-verified"
    elif "evc" in raw:
        return "Other EVC"
    elif "option expired" in raw:
        return "Option Expired (NA)"
    elif re.search(r'\b(?:not applicable|na)\b', raw, re.I):
        return "Not Applicable (NA)"
    elif "landing" in raw or "form selected" in raw or "draft" in raw or "profile" in raw:
        return "Not submitted"
    if re.search(r'^(?:fy|due\s*date|status|-+)[\s\-:]*$', raw, re.I):
        return "Not submitted"
    return "Not submitted"

NON_QUARTER_MONTHS = {
    "apr", "april", "may", "jul", "july", "aug", "august", 
    "oct", "october", "nov", "november", "jan", "january", "feb", "february"
}

def is_non_quarter_month(period_str: str) -> bool:
    p_lower = str(period_str or "").lower()
    return any(re.search(rf'\b{m}\b', p_lower) for m in NON_QUARTER_MONTHS)

SKELETON_NAME_REGEX = re.compile(
    r'^(?:taxpayer|client|user|individual|indicates\s*mandatory\s*fields|mandatory\s*fields|goods\s+and\s+services\s+tax|gst\s+common\s+portal|gst\s+portal|status|due\s*date|fy|financial\s*year|tax\s*period|return\s*period|filing\s*period|legal\s*name|trade\s*name|gstin|pan|na|-+)[\s\-:*]*$',
    re.I
)

SKELETON_PERIOD_REGEX = re.compile(
    r'^(?:due\s*date|status|fy|financial\s*year|tax\s*period|return\s*period|filing\s*period|na|-+)[\s\-:]*$',
    re.I
)

def normalize_form_type(raw_form: str) -> str:
    if not raw_form:
        return ""
    f = str(raw_form).strip()
    if re.search(r'\b(GSTR[-_ ]*1A)\b', f, re.I):
        return "GSTR-1A"
    if re.search(r'\b(GSTR[-_ ]*1(?:\s*/\s*IFF)?|IFF)\b', f, re.I):
        return "GSTR-1/IFF"
    if re.search(r'\b(GSTR[-_ ]*2A)\b', f, re.I):
        return "GSTR-2A"
    if re.search(r'\b(GSTR[-_ ]*2B)\b', f, re.I):
        return "GSTR-2B"
    if re.search(r'\b(GSTR[-_ ]*3B[Q]?)\b', f, re.I):
        return "GSTR-3B"
    if re.search(r'\b(CMP[-_ ]*08)\b', f, re.I):
        return "CMP-08"
    if re.search(r'\b(GSTR[-_ ]*4)\b', f, re.I):
        return "GSTR-4"
    if re.search(r'\b(GSTR[-_ ]*9C)\b', f, re.I):
        return "GSTR-9C"
    if re.search(r'\b(GSTR[-_ ]*9)\b', f, re.I):
        return "GSTR-9"
    if re.search(r'\b(GSTR[-_ ]*7)\b', f, re.I):
        return "GSTR-7"
    if re.search(r'\b(GSTR[-_ ]*8)\b', f, re.I):
        return "GSTR-8"
    if re.search(r'\b(ITR[-_ ]*[1-7])\b', f, re.I):
        m = re.search(r'\b(ITR[-_ ]*[1-7])\b', f, re.I)
        return m.group(1).upper().replace(" ", "-")
    return f

def normalize_period(raw_period: str) -> str:
    if not raw_period:
        return "Unknown Period"
    p = str(raw_period).strip()
    p = re.sub(r'\s+', ' ', p)
    # Reject skeleton placeholders
    if SKELETON_PERIOD_REGEX.search(p):
        return "Unknown Period"
    # Split before status keywords if multi-line text was captured from table cells
    p = re.split(r'\s+(?:Filed|Not Filed|To be Filed|Pending|Option expired|Due date)\b', p, flags=re.I)[0].strip()
    if not p or SKELETON_PERIOD_REGEX.search(p):
        return "Unknown Period"
    return p

def infer_portal(filing_type: str, gstin: str = "", portal_hint: str = "") -> str:
    """Infers the government compliance portal based on form type and identifiers."""
    ft = str(filing_type or "").upper()
    ph = str(portal_hint or "").lower()
    if "gst" in ph or ft.startswith("GSTR") or ft.startswith("CMP") or ft.startswith("ITC"):
        return "GST Portal"
    if "income tax" in ph or "itr" in ph or ft.startswith("ITR") or "FORM-10" in ft:
        return "Income Tax (ITD)"
    if "traces" in ph or "tds" in ph or any(k in ft for k in ["24Q", "26Q", "27Q", "27EQ"]):
        return "TRACES / TDS"
    if gstin and len(gstin) >= 15:
        return "GST Portal"
    if ft.startswith("ITR"):
        return "Income Tax (ITD)"
    return "GST Portal" if gstin else "Income Tax (ITD)"

def parse_date_str(d_str: str) -> date | None:
    """Safely parses common date formats into a date object."""
    if not d_str:
        return None
    d = str(d_str).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y", "%d-%b-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(d[:10], fmt).date()
        except Exception:
            pass
    return None

def compute_compliance_alert(status: str, due_date_str: str, last_updated_str: str) -> str:
    """Calculates due date aging and 30-day e-verification expiry status."""
    today = date.today()
    if status == "Submitted & E-verified":
        return "Complied"
    
    if status == "Submitted (e-verification pending)":
        up_date = parse_date_str(last_updated_str)
        if up_date:
            days_elapsed = (today - up_date).days
            if days_elapsed > 30:
                return f"🚨 Expired ({days_elapsed}d > 30d limit)"
            elif days_elapsed >= 20:
                return f"⚠️ Expiring Soon ({30 - days_elapsed}d left)"
            else:
                return f"Pending ({30 - days_elapsed}d left)"
        return "Pending e-Verification"

    if status == "Other EVC":
        return "EVC Action Required"

    if status in ("Option Expired (NA)", "Not Applicable (NA)"):
        return "Not Applicable"

    if status == "Not submitted":
        due_d = parse_date_str(due_date_str)
        if due_d:
            days_diff = (due_d - today).days
            if days_diff < 0:
                return f"⚠️ Overdue ({abs(days_diff)}d ago)"
            elif days_diff == 0:
                return "Due Today"
            elif days_diff <= 5:
                return f"Due Soon ({days_diff}d left)"
            else:
                return f"Due in {days_diff}d"
        return "Pending Filing"

    return "Status Pending"

def process_timelines():
    """Extracts, normalizes, and consolidates filings from sdc_session_timelines and tracker_dump."""
    print("Connecting to rawPayload.db to extract SDC Timelines...")
    try:
        conn = get_db_connection()
        cur = conn.cursor()
    except Exception as e:
        print(f"Failed to connect to database: {e}")
        return []
    
    try:
        # Order by start_time ASC so that newer sessions overwrite values for the same (PAN, GSTIN, Form, Filing Period)
        cur.execute("""
            SELECT session_id, pan, client_name, start_time, end_time, timeline_json
            FROM sdc_session_timelines
            ORDER BY start_time ASC
        """)
        rows = cur.fetchall()
    except Exception as e:
        print(f"Failed to read sdc_session_timelines: {e}")
        return []
    finally:
        conn.close()

    ltt_dict = {}

    for row in rows:
        session_id = row['session_id']
        timeline_json = row['timeline_json']
        if not timeline_json:
            continue
            
        try:
            timeline = json.loads(timeline_json)
        except Exception:
            continue
            
        if not timeline:
            continue
            
        start_url = timeline[0].get('url', '')
        start_time = timeline[0].get('timestamp', '')
        end_url = timeline[-1].get('url', '')
        end_time = timeline[-1].get('timestamp', '')
        site_history = f"Start: {start_url} ({start_time})\nEnd: {end_url} ({end_time})\nSteps: {len(timeline)}"
        
        # 1. Resolve session-wide identity (PAN, GSTIN, Names, Filing Preference)
        session_pan = row['pan'] or ""
        session_full_name = ""
        session_temp_name = ""
        session_gstin = ""
        session_filing_pref = ""
        
        for step in timeline:
            cap = step.get('captured_data') or {}
            pref = cap.get('filing_preference') or step.get('filing_preference')
            if pref and not session_filing_pref:
                session_filing_pref = str(pref).strip().title()
            if cap.get('pan'):
                session_pan = cap.get('pan')
            if cap.get('gstin'):
                session_gstin = cap.get('gstin')
            trade_n = cap.get('company_name') or cap.get('trade_name') or ''
            prop_n = cap.get('proprietor_name') or cap.get('legal_name') or cap.get('client_name') or ''
            c_name = trade_n if (trade_n and not SKELETON_NAME_REGEX.search(trade_n)) else prop_n
            if c_name and not SKELETON_NAME_REGEX.search(c_name):
                if not session_full_name or len(c_name) >= len(session_full_name):
                    session_full_name = c_name
            t_name = cap.get('client_temp_name') or ''
            if t_name and not SKELETON_NAME_REGEX.search(t_name):
                if not session_temp_name or len(t_name) >= len(session_temp_name):
                    session_temp_name = t_name
                
        row_cname = row['client_name'] or ""
        if SKELETON_NAME_REGEX.search(row_cname):
            row_cname = ""
        final_name = session_full_name or session_temp_name or row_cname
        if SKELETON_NAME_REGEX.search(final_name):
            final_name = ""
        if not session_pan and session_gstin and len(session_gstin) >= 12:
            session_pan = session_gstin[2:12]
            
        if not session_pan:
            continue

        # 2. Extract distinct filings from the timeline
        filings_in_session = {}
        for step in timeline:
            cap = step.get('captured_data')
            if not cap:
                continue
            raw_form = (cap.get('form') or cap.get('filing_type') or "").strip()
            norm_form = normalize_form_type(raw_form)
            if not norm_form or norm_form in ["Profile / Identity", "ITR (Landing / e-File)", "GST Return Filing", "Filing Dashboard"]:
                continue
            if "returns calendar" in norm_form.lower():
                continue

            raw_period = (cap.get('ay') or cap.get('period_label') or cap.get('period') or "").strip()
            norm_period = normalize_period(raw_period)
            status_raw = cap.get('status') or ""
            due_date = cap.get('due_date') or ""
            step_gstin = cap.get('gstin') or session_gstin
            step_time = step.get('timestamp') or end_time or start_time
            step_arn = cap.get('arn') or cap.get('ack_number') or cap.get('ack_no') or cap.get('arn_number') or step.get('arn') or ''
            
            step_pref = cap.get('filing_preference') or session_filing_pref or ""
            sub_key = (norm_form, norm_period)
            if sub_key not in filings_in_session:
                filings_in_session[sub_key] = {
                    "form": norm_form,
                    "period": norm_period,
                    "status_raw": status_raw,
                    "due_date": due_date,
                    "gstin": step_gstin,
                    "arn": step_arn,
                    "time": step_time,
                    "preference": step_pref
                }
            else:
                existing_f = filings_in_session[sub_key]
                if status_raw:
                    existing_f["status_raw"] = status_raw
                if due_date:
                    existing_f["due_date"] = due_date
                if step_gstin:
                    existing_f["gstin"] = step_gstin
                if step_arn:
                    existing_f["arn"] = step_arn
                if step_pref:
                    existing_f["preference"] = step_pref
                existing_f["time"] = step_time

        # Prune generic placeholder periods if a specific period exists for the same form
        forms_with_specific_period = {
            f_type for (f_type, p_label) in filings_in_session
            if p_label not in ("", "Current Period", "Unknown Period")
        }
        for (f_type, p_label) in list(filings_in_session.keys()):
            if p_label in ("Current Period", "Unknown Period") and f_type in forms_with_specific_period:
                del filings_in_session[(f_type, p_label)]
            elif p_label in ("", "Current Period", "Unknown Period"):
                f_entry = filings_in_session[(f_type, p_label)]
                if evaluate_status(f_entry.get("status_raw")) == "Not submitted" and not f_entry.get("due_date"):
                    del filings_in_session[(f_type, p_label)]

        if not filings_in_session:
            continue

        for (f_type, p_label), f_data in filings_in_session.items():
            filing_period = f_data["period"] if f_data["period"] else "Unknown Period"
            filing_type = f_data["form"]
            item_gstin = f_data["gstin"] or session_gstin
            submit_status = evaluate_status(f_data["status_raw"])
            item_pref = f_data.get("preference") or session_filing_pref or "Regular"
            if item_pref.lower() == "quarterly" and is_non_quarter_month(filing_period):
                if filing_type == "GSTR-3B" and submit_status == "Not submitted":
                    submit_status = "Not Applicable (NA)"
                elif filing_type in ("GSTR-1/IFF", "GSTR-1") and submit_status == "Not submitted":
                    submit_status = "Option Expired (NA)"

            arn_val = f_data.get("arn") or ""
            if arn_val and arn_val != "N/A" and submit_status == "Not submitted":
                submit_status = "Submitted & E-verified"
            
            portal_val = infer_portal(filing_type, item_gstin)
            key = (session_pan, item_gstin, filing_type, filing_period)
            
            if key not in ltt_dict:
                ltt_dict[key] = {
                    "PAN": session_pan,
                    "GSTIN": item_gstin,
                    "Client Name": final_name,
                    "Portal": portal_val,
                    "Filing Preference": item_pref,
                    "Filing Period": filing_period,
                    "Filing Type": filing_type,
                    "Submit Status": submit_status,
                    "ARN": arn_val,
                    "Due Date": f_data["due_date"],
                    "Session ID": session_id,
                    "Site History": site_history,
                    "Last Updated": f_data["time"],
                    "Compliance Alert": "",
                    "Discrepancy Note": ""
                }
            else:
                existing = ltt_dict[key]
                if f_data["status_raw"]:
                    existing["Submit Status"] = submit_status
                if f_data["due_date"]:
                    existing["Due Date"] = f_data["due_date"]
                if arn_val:
                    existing["ARN"] = arn_val
                existing["Session ID"] = session_id
                existing["Site History"] = site_history
                existing["Last Updated"] = f_data["time"]
                if session_full_name and (not existing["Client Name"] or existing["Client Name"] == session_temp_name):
                    existing["Client Name"] = session_full_name

    # 3. Merge entries from tracker_dump table so live and individually captured filings are always included
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, client_id, portal, period_label, arn_number, capture_method, status, raw_payload_json, created_at
            FROM tracker_dump
            ORDER BY id ASC
        """)
        td_rows = cur.fetchall()
        for r in td_rows:
            raw_json = r['raw_payload_json'] or '{}'
            try:
                p_data = json.loads(raw_json)
            except Exception:
                p_data = {}
            raw_p = p_data.get('raw_payload') or {}
            
            captures_to_process = raw_p.get('ltt_captures') or raw_p.get('assembler_captures')
            if not captures_to_process or not isinstance(captures_to_process, list):
                captures_to_process = [p_data]

            for item in captures_to_process:
                if not isinstance(item, dict):
                    continue
                t_pan = (item.get('pan') or p_data.get('pan') or raw_p.get('pan') or '').strip().upper()
                t_gstin = (item.get('gstin') or p_data.get('gstin') or raw_p.get('gstin') or '').strip().upper()
                if not t_pan and t_gstin and len(t_gstin) >= 12:
                    t_pan = t_gstin[2:12]
                if not t_pan:
                    continue

                t_name = (item.get('proprietor_name') or item.get('legal_name') or item.get('client_name') or 
                          item.get('name') or p_data.get('proprietor_name') or p_data.get('legal_name') or 
                          p_data.get('client_name') or raw_p.get('proprietor_name') or raw_p.get('client_name') or '')
                if SKELETON_NAME_REGEX.search(t_name):
                    t_name = ""
                raw_period_val = item.get('period_label') or item.get('period') or p_data.get('period_label') or r['period_label'] or ''
                t_period = normalize_period(raw_period_val)
                t_form = normalize_form_type(item.get('filing_type') or item.get('form') or p_data.get('filing_type') or r['portal'] or '')
                if not t_form or t_form in ["Profile / Identity", "ITR (Landing / e-File)", "GST Return Filing", "Filing Dashboard"]:
                    continue
                if "returns calendar" in t_form.lower():
                    continue
                t_status = evaluate_status(item.get('status') or r['status'] or p_data.get('status') or '')
                if t_period in ("Current Period", "Unknown Period", "Due Date -", "Status -", "FY -") and t_status in ("Not submitted", "Initiated", "FY -"):
                    continue
                t_due = item.get('due_date') or p_data.get('due_date') or ''
                t_arn = (item.get('arn') or item.get('arn_number') or item.get('ack_number') or 
                         p_data.get('arn_number') or r['arn_number'] or '').strip()
                if t_arn and t_arn != "N/A" and t_status == "Not submitted":
                    t_status = "Submitted & E-verified"
                
                t_pref = (item.get('filing_preference') or 
                          p_data.get('filing_preference') or 
                          raw_p.get('filing_preference') or 
                          (item.get('scraped_data') or {}).get('filing_preference') or '').strip().title()
                if (t_pref.lower() == "quarterly" or "quarterly" in str(item.get("scraped_data", {})).lower()) and is_non_quarter_month(t_period):
                    if t_form == "GSTR-3B" and t_status == "Not submitted":
                        t_status = "Not Applicable (NA)"
                    elif t_form in ("GSTR-1/IFF", "GSTR-1") and t_status == "Not submitted":
                        t_status = "Option Expired (NA)"

                portal_val = infer_portal(t_form, t_gstin, portal_hint=r['portal'])
                t_key = (t_pan, t_gstin, t_form, t_period)

                if t_key not in ltt_dict:
                    ltt_dict[t_key] = {
                        "PAN": t_pan,
                        "GSTIN": t_gstin,
                        "Client Name": t_name,
                        "Portal": portal_val,
                        "Filing Preference": t_pref or "Regular",
                        "Filing Period": t_period if t_period else "Unknown Period",
                        "Filing Type": t_form,
                        "Submit Status": t_status,
                        "ARN": t_arn if t_arn != "N/A" else "",
                        "Due Date": t_due,
                        "Session ID": item.get('session_id') or p_data.get('session_id') or f"TD-{r['id']}",
                        "Site History": f"Captured via {item.get('capture_method') or r['capture_method']}",
                        "Last Updated": item.get('last_viewed_at') or item.get('updated_at') or r['created_at'],
                        "Compliance Alert": "",
                        "Discrepancy Note": ""
                    }
                else:
                    existing = ltt_dict[t_key]
                    if t_status and t_status != "Not submitted":
                        existing["Submit Status"] = t_status
                    if t_pref and (not existing.get("Filing Preference") or existing.get("Filing Preference") == "Regular"):
                        existing["Filing Preference"] = t_pref
                    if t_due:
                        existing["Due Date"] = t_due
                    if t_arn and t_arn != "N/A":
                        existing["ARN"] = t_arn
                    if t_name and (not existing["Client Name"] or SKELETON_NAME_REGEX.search(existing["Client Name"])):
                        existing["Client Name"] = t_name
        conn.close()
    except Exception as e:
        print(f"Failed to merge tracker_dump into LTT: {e}")

    # 4. Post-processing: Compute Compliance / Aging Alerts & Cross-form Discrepancies
    # Pass 1: Propagate Quarterly preference and resolve non-quarterly GSTR-3B
    for key, item in ltt_dict.items():
        pan, gstin, f_type, period = key
        pref = str(item.get("Filing Preference", "")).lower()
        if f_type == "GSTR-3B" and item["Submit Status"] == "Not submitted" and is_non_quarter_month(period):
            comp_key = (pan, gstin, "GSTR-1/IFF", period)
            comp_item = ltt_dict.get(comp_key)
            if pref == "quarterly" or (comp_item and comp_item.get("Submit Status") in ("Option Expired (NA)", "Not Applicable (NA)")):
                item["Submit Status"] = "Not Applicable (NA)"
                item["Filing Preference"] = "Quarterly"

    # Pass 2: Compute individual aging/compliance alert
    gst_period_forms = {}
    for key, item in ltt_dict.items():
        pan, gstin, f_type, period = key
        item["Compliance Alert"] = compute_compliance_alert(
            item["Submit Status"], item.get("Due Date", ""), item.get("Last Updated", "")
        )
        if item.get("Portal") == "GST Portal":
            g_key = (pan, gstin, period)
            if g_key not in gst_period_forms:
                gst_period_forms[g_key] = {}
            gst_period_forms[g_key][f_type] = item["Submit Status"]

    # Check for GSTR-1 vs GSTR-3B discrepancies (ONLY for regular monthly filers or quarter-ending months)
    for (pan, gstin, period), forms in gst_period_forms.items():
        if is_non_quarter_month(period):
            continue
        g1_status = forms.get("GSTR-1/IFF") or forms.get("GSTR-1")
        g3b_status = forms.get("GSTR-3B")
        
        g1_filed = g1_status in ("Submitted & E-verified", "Submitted (e-verification pending)")
        g3b_filed = g3b_status in ("Submitted & E-verified", "Submitted (e-verification pending)")

        if g1_filed and g3b_status == "Not submitted":
            k = (pan, gstin, "GSTR-3B", period)
            if k in ltt_dict:
                ltt_dict[k]["Discrepancy Note"] = "⚠️ GSTR-3B Pending (GSTR-1 Filed)"
        elif g3b_filed and g1_status == "Not submitted":
            k = (pan, gstin, "GSTR-1/IFF", period)
            if k in ltt_dict:
                ltt_dict[k]["Discrepancy Note"] = "⚠️ GSTR-1 Pending (GSTR-3B Filed)"

    # Pass 3: Consolidate Income Tax (ITR) to strictly ONE final entry registered by the final payload
    # GST filings and other portals remain completely untouched.
    final_ltt_list = []
    itr_candidates_by_pan = {}

    for item in ltt_dict.values():
        if item.get("Portal") == "Income Tax (ITD)":
            pan = item.get("PAN")
            if pan:
                if pan not in itr_candidates_by_pan:
                    itr_candidates_by_pan[pan] = []
                itr_candidates_by_pan[pan].append(item)
            else:
                final_ltt_list.append(item)
        else:
            # GST and TRACES / TDS: completely untouched
            final_ltt_list.append(item)

    def _itr_final_payload_key(r):
        """Returns a composite comparison key to determine the final, authoritative payload."""
        ts = str(r.get("Last Updated") or "")
        form_val = str(r.get("Filing Type") or "")
        form_rank = 2 if re.match(r"^ITR-[1-7]", form_val, re.I) else (1 if "ITR" in form_val else 0)
        period_val = str(r.get("Filing Period") or "")
        period_rank = 0
        if period_val.startswith("AY 20"):
            try:
                period_rank = int(period_val.split()[1][:4])
            except Exception:
                period_rank = 2000
        elif period_val not in ("", "Unknown Period", "Current Period"):
            period_rank = 1000
        arn_val = str(r.get("ARN") or "").strip()
        has_arn = 1 if (arn_val and arn_val not in ("-", "N/A", "None")) else 0
        st_val = str(r.get("Submit Status") or "").lower()
        st_rank = 2 if "verified" in st_val else (1 if "pending" in st_val else 0)

        return (ts, form_rank, period_rank, has_arn, st_rank)

    for pan, candidate_list in itr_candidates_by_pan.items():
        # Pick the single final entry registered by the final payload
        candidate_list.sort(key=_itr_final_payload_key, reverse=True)
        final_ltt_list.append(candidate_list[0])

    return final_ltt_list

def get_ltt_dataset():
    """Authoritative API returning the sorted LTT dataset and computed KPI metrics."""
    data = process_timelines()
    if not data:
        return [], {}

    # Sort descending by Last Updated
    def sort_key(x):
        return str(x.get("Last Updated") or "")
    data = sorted(data, key=sort_key, reverse=True)

    # Compute KPI summary metrics
    total = len(data)
    verified = sum(1 for d in data if d.get("Submit Status") == "Submitted & E-verified")
    pending = sum(1 for d in data if d.get("Submit Status") == "Submitted (e-verification pending)")
    other_evc = sum(1 for d in data if d.get("Submit Status") == "Other EVC")
    not_sub = sum(1 for d in data if d.get("Submit Status") == "Not submitted")
    overdue = sum(1 for d in data if "Overdue" in d.get("Compliance Alert", ""))
    
    defaulters = sum(
        1 for d in data if (
            ("Overdue" in d.get("Compliance Alert", "") or 
             "Expired" in d.get("Compliance Alert", "") or 
             "Expiring Soon" in d.get("Compliance Alert", "") or 
             bool(d.get("Discrepancy Note"))) and
            d.get("Submit Status") not in ("Option Expired (NA)", "Not Applicable (NA)", "Submitted & E-verified")
        )
    )

    itd_total = sum(1 for d in data if d.get("Portal") == "Income Tax (ITD)")
    itd_verif = sum(1 for d in data if d.get("Portal") == "Income Tax (ITD)" and d.get("Submit Status") == "Submitted & E-verified")
    gst_total = sum(1 for d in data if d.get("Portal") == "GST Portal")
    gst_verif = sum(1 for d in data if d.get("Portal") == "GST Portal" and d.get("Submit Status") == "Submitted & E-verified")

    applicable_total = sum(1 for d in data if d.get("Submit Status") not in ("Option Expired (NA)", "Not Applicable (NA)"))
    compliance_rate = round((verified / applicable_total * 100), 1) if applicable_total > 0 else 0.0

    kpis = {
        "total_filings": total,
        "verified_count": verified,
        "pending_verif_count": pending,
        "other_evc_count": other_evc,
        "not_submitted_count": not_sub,
        "overdue_count": overdue,
        "defaulters_count": defaulters,
        "compliance_rate": compliance_rate,
        "itd_total": itd_total,
        "itd_verified": itd_verif,
        "gst_total": gst_total,
        "gst_verified": gst_verif
    }

    return data, kpis

def generate_ltt_excel():
    """
    Generates a rich, multi-sheet, WCAG-compliant Excel workbook:
    - Sheet 1: Executive Summary (KPIs & Compliance Health)
    - Sheet 2: Master LTT (All filings, with ARNs, Alerts, and Auto-filters)
    - Sheet 3: Income Tax Filings
    - Sheet 4: GST Filings
    - Sheet 5: Action Required & Defaulters
    - Sheet 6: GST Compliance Matrix
    """
    data, kpis = get_ltt_dataset()
    if not data:
        print("No SDC timeline data found or extracted.")
        return None

    df_master = pd.DataFrame(data)
    cols = [
        "PAN", "GSTIN", "Client Name", "Portal", "Filing Preference", "Filing Period", "Filing Type",
        "Submit Status", "ARN", "Due Date", "Compliance Alert", "Discrepancy Note",
        "Session ID", "Last Updated", "Site History"
    ]
    for c in cols:
        if c not in df_master.columns:
            df_master[c] = ""
    df_master = df_master[cols]

    live_dir = os.path.join(os.path.expanduser("~"), "AmanAssociates_Sera")
    os.makedirs(live_dir, exist_ok=True)
    output_file = os.path.join(live_dir, "Live_Tracking_Table_LTT.xlsx")

    # Styling constants
    font_family = "Calibri"
    cell_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    header_border = Border(
        left=Side(style='thin', color='203764'),
        right=Side(style='thin', color='203764'),
        top=Side(style='medium', color='203764'),
        bottom=Side(style='medium', color='203764')
    )
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name=font_family, size=11)

    style_verified = {
        "fill": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "font": Font(color="145A32", bold=True, name=font_family, size=10)
    }
    style_pending = {
        "fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "font": Font(color="7D6608", bold=True, name=font_family, size=10)
    }
    style_not_submitted = {
        "fill": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "font": Font(color="78281F", bold=True, name=font_family, size=10)
    }
    style_expired = {
        "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
        "font": Font(color="595959", bold=True, name=font_family, size=10)
    }
    style_evc = {
        "fill": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        "font": Font(color="1B4F72", bold=True, name=font_family, size=10)
    }
    style_alert_danger = {
        "fill": PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid"),
        "font": Font(color="78281F", bold=True, name=font_family, size=10)
    }
    style_alert_warning = {
        "fill": PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid"),
        "font": Font(color="7D6608", bold=True, name=font_family, size=10)
    }

    def style_standard_table(ws, has_status=True, status_col_name="Submit Status"):
        ws.views.sheetView[0].showGridLines = True
        ws.freeze_panes = "A2"

        # Apply header styling
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = header_border

        status_col = None
        alert_col = None
        for col_idx in range(1, ws.max_column + 1):
            h_val = str(ws.cell(row=1, column=col_idx).value or '').strip().lower()
            if h_val == status_col_name.lower():
                status_col = col_idx
            elif h_val in ("compliance alert", "alert"):
                alert_col = col_idx

        for row_idx in range(2, ws.max_row + 1):
            for c_idx in range(1, ws.max_column + 1):
                c = ws.cell(row=row_idx, column=c_idx)
                c.border = cell_border
                c.font = Font(name=font_family, size=10)
                c.alignment = Alignment(vertical="center")

            # Color Submit Status
            if status_col:
                s_cell = ws.cell(row=row_idx, column=status_col)
                s_val = str(s_cell.value or '').strip().lower()
                target_style = None
                if "verified" in s_val or "filed" in s_val:
                    target_style = style_verified
                elif "pending" in s_val:
                    target_style = style_pending
                elif "not submitted" in s_val or "unfiled" in s_val:
                    target_style = style_not_submitted
                elif "expired" in s_val or "na" in s_val:
                    target_style = style_expired
                elif "evc" in s_val:
                    target_style = style_evc

                if target_style:
                    s_cell.fill = target_style["fill"]
                    s_cell.font = target_style["font"]
                    s_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Color Compliance Alert
            if alert_col:
                a_cell = ws.cell(row=row_idx, column=alert_col)
                a_val = str(a_cell.value or '').strip()
                if "Overdue" in a_val or "Expired" in a_val or "🚨" in a_val:
                    a_cell.fill = style_alert_danger["fill"]
                    a_cell.font = style_alert_danger["font"]
                elif "Soon" in a_val or "⚠️" in a_val or "Due in" in a_val:
                    a_cell.fill = style_alert_warning["fill"]
                    a_cell.font = style_alert_warning["font"]
                elif "Complied" in a_val:
                    a_cell.fill = style_verified["fill"]
                    a_cell.font = style_verified["font"]
                a_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-adjust column widths with bounds
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                max_len = max(max_len, len(val.split('\n')[0]))
            ws.column_dimensions[col_letter].width = max(12, min(max_len + 4, 60))

        # Enable auto-filter
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    # Build Sub-dataframes
    df_itd = df_master[df_master['Portal'] == 'Income Tax (ITD)'].copy()
    itd_cols = ["PAN", "Client Name", "Filing Period", "Filing Type", "Submit Status", "ARN", "Due Date", "Compliance Alert", "Last Updated"]
    df_itd = df_itd[[c for c in itd_cols if c in df_itd.columns]]

    df_gst = df_master[df_master['Portal'] == 'GST Portal'].copy()
    gst_cols = ["GSTIN", "PAN", "Client Name", "Filing Period", "Filing Type", "Submit Status", "ARN", "Due Date", "Compliance Alert", "Discrepancy Note", "Last Updated"]
    df_gst = df_gst[[c for c in gst_cols if c in df_gst.columns]]

    df_action = df_master[
        (
            df_master['Compliance Alert'].str.contains("Overdue|Expired|Soon|🚨|⚠️", case=False, na=False) |
            (df_master['Discrepancy Note'] != "") |
            (df_master['Submit Status'] == "Not submitted")
        ) &
        (~df_master['Submit Status'].isin(["Option Expired (NA)", "Not Applicable (NA)", "Submitted & E-verified"]))
    ].copy()

    # Build GST Compliance Matrix (Clients vs Periods)
    matrix_rows = []
    gst_clients = df_gst[['GSTIN', 'Client Name']].drop_duplicates()
    gst_periods = sorted(list(set(df_gst['Filing Period'].dropna())), reverse=True)
    
    for _, cl in gst_clients.iterrows():
        gstin = cl['GSTIN']
        cname = cl['Client Name']
        row_dict = {"GSTIN": gstin, "Client Name": cname}
        for p in gst_periods:
            sub = df_gst[(df_gst['GSTIN'] == gstin) & (df_gst['Filing Period'] == p)]
            status_parts = []
            for _, r in sub.iterrows():
                f_short = "G1" if "1" in r['Filing Type'] else ("3B" if "3B" in r['Filing Type'] else r['Filing Type'])
                st_l = str(r['Submit Status']).lower()
                if "verified" in st_l:
                    s_icon = "✓"
                elif "pending" in st_l:
                    s_icon = "⏳"
                elif "na" in st_l or "expired" in st_l or "not applicable" in st_l:
                    s_icon = "NA"
                else:
                    s_icon = "✗"
                status_parts.append(f"{f_short}:{s_icon}")
            row_dict[p] = " | ".join(status_parts) if status_parts else "-"
        matrix_rows.append(row_dict)
    df_matrix = pd.DataFrame(matrix_rows) if matrix_rows else pd.DataFrame(columns=["GSTIN", "Client Name"])

    target_out = output_file
    try:
        writer = pd.ExcelWriter(target_out, engine='openpyxl')
    except PermissionError:
        target_out = os.path.join(live_dir, f"Live_Tracking_Table_LTT_{datetime.now().strftime('%H%M%S')}.xlsx")
        writer = pd.ExcelWriter(target_out, engine='openpyxl')

    # 1. Executive Summary Sheet
    wb = writer.book
    ws_summary = wb.create_sheet(title='Executive Summary', index=0)
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Title Banner
    ws_summary.merge_cells('A1:F1')
    title_cell = ws_summary['A1']
    title_cell.value = "Live Tracking Table (LTT) - Compliance Executive Dashboard"
    title_cell.font = Font(name=font_family, size=15, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="0D233A", end_color="0D233A", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 36

    # Generated timestamp
    ws_summary.merge_cells('A2:F2')
    sub_cell = ws_summary['A2']
    sub_cell.value = f"Generated on: {datetime.now().strftime('%d %b %Y, %I:%M %p')} | Total Filings Tracked: {kpis['total_filings']}"
    sub_cell.font = Font(name=font_family, size=10, italic=True, color="595959")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[2].height = 20

    # Summary KPI Blocks
    kpi_headers = [
        ("Total Filings", kpis['total_filings'], "1F4E79"),
        ("Submitted & Verified", kpis['verified_count'], "1E8449"),
        ("e-Verif Pending", kpis['pending_verif_count'], "B7950B"),
        ("Overdue / Unfiled", kpis['not_submitted_count'], "A93226"),
        ("Overall Compliance", f"{kpis['compliance_rate']}%", "117A65")
    ]
    for idx, (title, val, hex_c) in enumerate(kpi_headers):
        c_letter = get_column_letter(idx + 1)
        ws_summary[f'{c_letter}4'] = title
        ws_summary[f'{c_letter}4'].font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
        ws_summary[f'{c_letter}4'].fill = PatternFill(start_color=hex_c, end_color=hex_c, fill_type="solid")
        ws_summary[f'{c_letter}4'].alignment = Alignment(horizontal="center", vertical="center")

        ws_summary[f'{c_letter}5'] = val
        ws_summary[f'{c_letter}5'].font = Font(name=font_family, size=14, bold=True, color=hex_c)
        ws_summary[f'{c_letter}5'].alignment = Alignment(horizontal="center", vertical="center")
        ws_summary[f'{c_letter}5'].border = cell_border
        ws_summary.column_dimensions[c_letter].width = 22

    # Portal Breakdown Table
    ws_summary['A7'] = "Portal Compliance Breakdown"
    ws_summary['A7'].font = Font(name=font_family, size=12, bold=True, color="1F4E79")
    
    breakdown_data = [
        ["Portal Jurisdiction", "Total Tracked", "Submitted & Verified", "Compliance Rate"],
        ["Income Tax Department (ITD)", kpis['itd_total'], kpis['itd_verified'], f"{round((kpis['itd_verified']/kpis['itd_total']*100),1) if kpis['itd_total']>0 else 0.0}%"],
        ["GST Common Portal", kpis['gst_total'], kpis['gst_verified'], f"{round((kpis['gst_verified']/kpis['gst_total']*100),1) if kpis['gst_total']>0 else 0.0}%"]
    ]
    for r_idx, row in enumerate(breakdown_data, start=8):
        for c_idx, val in enumerate(row, start=1):
            cell = ws_summary.cell(row=r_idx, column=c_idx, value=val)
            cell.border = cell_border
            if r_idx == 8:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.font = Font(name=font_family, size=10, bold=(c_idx==1))
                cell.alignment = Alignment(horizontal=("left" if c_idx==1 else "center"), vertical="center")

    # Write other sheets
    df_master.to_excel(writer, index=False, sheet_name='Master LTT')
    style_standard_table(writer.sheets['Master LTT'])

    if not df_itd.empty:
        df_itd.to_excel(writer, index=False, sheet_name='Income Tax Filings')
        style_standard_table(writer.sheets['Income Tax Filings'])

    if not df_gst.empty:
        df_gst.to_excel(writer, index=False, sheet_name='GST Filings')
        style_standard_table(writer.sheets['GST Filings'])

    if not df_action.empty:
        df_action.to_excel(writer, index=False, sheet_name='Action Required & Defaulters')
        style_standard_table(writer.sheets['Action Required & Defaulters'])

    if not df_matrix.empty:
        df_matrix.to_excel(writer, index=False, sheet_name='GST Compliance Matrix')
        style_standard_table(writer.sheets['GST Compliance Matrix'], has_status=False)

    writer.close()
    print(f"Success! Generated multi-sheet LTT report at: {target_out} ({len(data)} records).")

    # Also automatically refresh the live CSV data feeds and linked workbook
    try:
        export_ltt_live_feed(output_dir=live_dir)
    except Exception as e:
        if DEBUG:
            print(f"[SDC_Parser] Notice updating live feeds during excel export: {e}")

    return target_out


def export_ltt_live_feed(output_dir=None, force_recreate_workbook=False):
    """
    Exports clean, high-performance CSV data feeds for Microsoft Excel Power Query / Data Connections.
    
    Generates:
    - LTT_Data_Feed.csv: Full live tabular dataset (Master LTT) formatted for 1-click Excel refresh.
    - LTT_Defaulters_Feed.csv: Filtered action-required / defaulters feed.
    - Live_Tracking_Table_Live.xlsx: The live-linked Excel workbook pre-wired with Excel's native QueryTable / PowerQuery.
      If Live_Tracking_Table_Live.xlsx already exists, it is NEVER overwritten, ensuring user formulas,
      notes, custom columns, and conditional formatting are permanently preserved!
    """
    import subprocess
    if output_dir is None:
        output_dir = os.path.join(os.path.expanduser("~"), "AmanAssociates_Sera")
    os.makedirs(output_dir, exist_ok=True)

    data, kpis = get_ltt_dataset()
    if not data:
        print("[SDC_Parser] Notice: No LTT data to export to live feed.")
        return None, None

    df_master = pd.DataFrame(data)
    
    # Clean up multi-line string columns to ensure exactly one line per row in CSV/Excel
    for col in df_master.select_dtypes(include=['object']):
        df_master[col] = (
            df_master[col]
            .astype(str)
            .replace('None', '')
            .replace('nan', '')
            .str.replace('\r\n', ' ; ')
            .str.replace('\n', ' ; ')
            .str.replace('\r', ' ; ')
        )

    # 1. Export Master CSV Feed
    csv_master_path = os.path.join(output_dir, "LTT_Data_Feed.csv")
    df_master.to_csv(csv_master_path, index=False, encoding='utf-8-sig')

    # Also keep a copy in local SDC_Parser dir for local access
    local_csv = os.path.join(os.path.dirname(os.path.abspath(__file__)), "LTT_Data_Feed.csv")
    try:
        df_master.to_csv(local_csv, index=False, encoding='utf-8-sig')
    except Exception:
        pass

    # 2. Export Defaulters CSV Feed
    df_action = df_master[
        (
            df_master['Compliance Alert'].str.contains("Overdue|Expired|Soon|🚨|⚠️", case=False, na=False) |
            (df_master['Discrepancy Note'] != "") |
            (df_master['Submit Status'] == "Not submitted")
        ) &
        (~df_master['Submit Status'].isin(["Option Expired (NA)", "Not Applicable (NA)", "Submitted & E-verified"]))
    ].copy()
    csv_action_path = os.path.join(output_dir, "LTT_Defaulters_Feed.csv")
    df_action.to_csv(csv_action_path, index=False, encoding='utf-8-sig')

    # 3. Create or maintain the Live Excel Workbook (Power Query / Data Connection)
    live_xlsx_path = os.path.join(output_dir, "Live_Tracking_Table_Live.xlsx")
    ps1_script = os.path.join(os.path.dirname(os.path.abspath(__file__)), "create_live_workbook.ps1")
    if os.path.exists(ps1_script):
        try:
            cmd = [
                "powershell", "-ExecutionPolicy", "Bypass", "-File", ps1_script,
                "-csvPath", csv_master_path,
                "-xlsxPath", live_xlsx_path
            ]
            if force_recreate_workbook:
                cmd.append("-force")
            subprocess.run(cmd, capture_output=True, text=True, timeout=20)
        except Exception as e:
            if DEBUG:
                print(f"[SDC_Parser] PowerShell workbook creation notice: {e}")

    print(f"[SDC_Parser] Live data feed updated: {csv_master_path} ({len(df_master)} records)")
    return csv_master_path, live_xlsx_path


if __name__ == '__main__':
    print("--- SDC Parser (Live Tracking Table) ---")
    generate_ltt_excel()
    export_ltt_live_feed()
